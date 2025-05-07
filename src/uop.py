##
## UOP v1.0
##
## Installation:
##  
## 1) Virtual Environment
## Create a virtual environment
## python3 -m venv uopc_virtual_env
## 
## 2) Activate your virtual environment
## Linux/Mac
## source uopc_virtual_env/bin/activate
## Example:
## 
## $ source uopc_virtual_env/bin/activate
## (uopc_virtual_env) $
## Deactivate virtual environment
## deactivate
## Example:
## 
## (uopc_virtual_env) $ deactivate
## 
## 3) Install dependencies
## Install dependencies from requirements.txt
## pip install -r requirements.txt
## 
## 4) Run python uop.py
## python uop.py



## Libraries ##################################################################

# Standard Library
import argparse
import colorlog
import folium
import logging
import loguru
import os
import random
import sqlite3
import sys
import urllib.parse 
from datetime import datetime
from pathlib import Path

# Third-Party Libraries
from fastapi import FastAPI, HTTPException
from fastapi.responses import HTMLResponse, JSONResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from folium import Map, plugins
from folium.plugins import MarkerCluster
# The Haversine formula to measure the straight-line distance between two geographical points (latitude and longitude) on the surface of the Earth
from haversine import haversine,Unit  
import numpy as np
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import pandas as pd
import tabulate
import uvicorn


## Initialize the FastAPI app #################################################
app = FastAPI()


## Serve static files in html fodler 
##app.mount("/html", StaticFiles(directory="html"), name="html")



## Arguments ##################################################################
# Define command-line arguments
parser = argparse.ArgumentParser(description="UPF Optimal Placer v1.0")
parser.add_argument("--host", type=str, default="127.0.0.1", help="Host address")
parser.add_argument("--port", type=int, default=8181, help="Port number")
parser.add_argument("--reload", action="store_true", default=True, help="Enable auto-reload")
parser.add_argument("--log-level", type=str, default="debug", choices=["debug", "info", "warning", "error", "critical"], help="Logging level")

# Parse arguments
args = parser.parse_args()



## CONSTANTS ##################################################################

# Path and database file name 
DATABASE_DIR = "./database/"  # Path to the directory
DATABASE_NAME = "network.db"  # Name of the database file

# Combine the directory and filename to create the full path
DATABASE_PATH = os.path.join(DATABASE_DIR, DATABASE_NAME)


# Base directory where the files are generated
current_path = os.getcwd()
BASE_DIR = os.path.join(os.getcwd(), 'output')  # Apunta a la carpeta 'output' en el directorio actual



## Loggin #####################################################################

# Define log file path (adjust as needed)
LOG_DIR = "logs"
LOG_FILE = os.path.join(LOG_DIR, f"uop_server.log")

# Ensure log directory exists
os.makedirs(LOG_DIR, exist_ok=True)

# Color definition
log_colors = {
    "DEBUG": "cyan",  # Azul
    "INFO": "green",   # Verde
    "WARNING": "yellow", # Amarillo
    "ERROR": "red",    # Rojo
    "CRITICAL": "bold_red", # Magenta
}

# Log formatter 
log_formatter = colorlog.ColoredFormatter(
    "%(log_color)s%(levelname)-8s%(reset)s %(message)s",  
    log_colors=log_colors,
)

# Console handler (prints logs to the screen)
console_handler = logging.StreamHandler()
console_handler.setFormatter(log_formatter)

# Configure logger
# Convert log level string to logging module constant
MY_LOG_LEVEL = getattr(logging, args.log_level.upper(), logging.DEBUG)

logger = logging.getLogger()
logger.setLevel(MY_LOG_LEVEL) 
logger.addHandler(console_handler)




## Database ###################################################################

# Database view to be used
view_name = "UOP_1_Shared_UPF_Edge_Far_Edge_DCs_location_info_v3"

# Views to query to generate tables for Word doc
views = [
    "Table_II_UOP_VIEW_0_Nearby_Cell_Sites_Far_Edge_DCs",
    "Table_III_UOP_VIEW_1_WITH_EDCs_FOR_A_SHARED_UPF",
    "Table_IV_UOP_VIEW_2_WITH_EDCs_FOR_A_SHARED_UPF_with_NSSAI",
    "Table_V_UOP_VIEW_3_WITH_EDCs_FOR_A_DEDICATED_UPF",
    "Table_VI_UOP_VIEW_4A_WITH_FE_EDCs_FOR_A_SHARED_UPF"
]

# Excel sheets names
sheet_names = {
    "Table_II_UOP_VIEW_0_Nearby_Cell_Sites_Far_Edge_DCs": "Table_II",
    "Table_III_UOP_VIEW_1_WITH_EDCs_FOR_A_SHARED_UPF": "Table_III",
    "Table_IV_UOP_VIEW_2_WITH_EDCs_FOR_A_SHARED_UPF_with_NSSAI": "Table_IV",
    "Table_V_UOP_VIEW_3_WITH_EDCs_FOR_A_DEDICATED_UPF": "Table_V",
    "Table_VI_UOP_VIEW_4A_WITH_FE_EDCs_FOR_A_SHARED_UPF": "Table_VI"
}



## Excel output  ##############################################################

# Output directory
output_excel = "output"  
os.makedirs(output_excel, exist_ok=True)

# HTML output directory
output_html = "html"  
os.makedirs(output_html, exist_ok=True)



## Map info ###################################################################

# Coordinates of your position
my_location = (32.841362, -96.784582) # SMU Dallas Campus

#map_center = (32.78130935199716, -96.81876935286627)  # New center for the map

arlington_coords = (32.73821052933643, -97.11275695630508) # Arlington National DC
utsw_coords  = (32.817195, -96.843869) ## UTSW Medical Center Dallas // W1225536259

# File static html
my_html_output2_name = 'uopv1-map'
timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
my_html_output2 = os.path.join(output_html, f"{my_html_output2_name}_{timestamp}.html")

all_data_html_1 = ""
all_data_html_2 = ""
all_data_html_3 = ""
all_data_html_4 = ""



## Some variables #############################################################
my_closest_fard_edge_is = ()
my_closest_fard_edge_is2 = ()

my_closest_fard_edge_name_is2 = None
my_closest_edge_name_is2 = None
my_closest_fard_edge_name_is3 = None
my_closest_edge_name_is3 = None



#############################################
## To use or to not use fixed values in: CPU, Mem, Disk, IOPS, BW
use_random_values = 1
#############################################

# CPU variables ------------------------------------------------
#cpu_min_value = 10.0
#cpu_max_value = 95.0
#cpu_min_thresold_value = 20.0
#cpu_max_thresold_value = 90.0
cpu_min_value_cnf_deployed = 70
cpu_max_value_cnf_deployed = 95
cpu_min_value_cnf_not_deployed = 10
cpu_max_value_cnf_not_deployed = 50

# Memory variables ---------------------------------------------
#mem_min_value = 10.0
#mem_max_value = 95.0
#mem_min_thresold_value = 0.10
#mem_max_thresold_value = 0.70
mem_min_value_cnf_deployed = 0.10
mem_max_value_cnf_deployed = 0.30
mem_min_value_cnf_not_deployed = 0.70
mem_max_value_cnf_not_deployed = 0.85

# Disk variables -----------------------------------------------
#disk_min_value = 1
disk_max_value = 4
disk_max_value2 = 4.000
#disk_min_thresold_value = 5.0
#disk_max_thresold_value = 75.0
disk_min_value_cnf_deployed = 0.10
disk_max_value_cnf_deployed = 0.40
disk_min_value_cnf_not_deployed = 0.70
disk_max_value_cnf_not_deployed = 0.85

# IOPS (Input/Output write/read per second)
#Range: 
# 100,000–1,500,000 IOPS
iops_min_value = 100000
iops_max_value = 1500000

# Bandwith -----------------------------------------------------
#Range: 
# 1000-25000Mbps 
# 1-25Gbps
# BW: Gpbs
bw_min_value = 1.00
bw_max_value = 25.00

# Latency ------------------------------------------------------
#latency_min_value = 0.00
#latency_max_value = 0.1
latency_min_value = 0.0
latency_max_value = 1000.0


# Weight -------------------------------------------------------
cpu_weight = 1
ram_weight = 1
disk_weight = 0.25
iops_weight = 0.5
bw_weight = 1
distance_weight = 1
latency_weight = 1

#manufacturer_weight = 1
#platform_weight = 1
#deployed_cnf_weight = 1


# Div (normalize)  
cpu_div  	 = 100
#cpu_div  	 = cpu_max_value 
ram_div   	 = 100
#ram_div   	 = mem_max_value
disk_div 	 = 100
#disk_div 	 = disk_max_value 
#iops_div   	 = 100
iops_div     = iops_max_value
#bw_div   	 = 1000
bw_div   	 = bw_max_value
distance_div = 100 
latency_div  = 100



## ++++++++++++++++++++++++++++++++++++++++++++++++++++++
## | Aux Functions                                      | 
## ++++++++++++++++++++++++++++++++++++++++++++++++++++++

def print_table(data, role_name,headers):
    """Print tables in tabular format"""
    table_data = [[d[col] for col in headers] for d in data]

    print(f"\nFinal Table for Role: {role_name} (Interleaved Scores + Total Score):")
    print(tabulate.tabulate(table_data, headers=headers, tablefmt="grid"))



def get_best_edge_server(data):
    """Select the best server in the list based on total_score, cpu_usage, and cpu_score + ram_score."""
    best_server = None

    for server in data:
        if best_server is None:
            best_server = server
        else:
            if server['total_score'] > best_server['total_score']:
                best_server = server
            elif server['total_score'] == best_server['total_score']:
                # Si tienen el mismo total_score, priorizar mayor cpu_usage
                if (server['cpu_score'] + server['ram_score']) > (best_server['cpu_score'] + best_server['ram_score']):
                    best_server = server

    return [best_server] if best_server else []



def get_random_value_in_db():
    """Random values for CNF, CPU, Mem, Disk, BW, IOPS."""

    global my_closest_fard_edge_name_is2
    global my_closest_edge_name_is2
    global my_closest_fard_edge_name_is3
    global my_closest_edge_name_is3

    # Connect to the SQLite database
    loguru.logger.debug(f"Trying to connect to database ...")
    conn = sqlite3.connect(DATABASE_PATH)
    cursor = conn.cursor()
    loguru.logger.debug(f"Trying to connect to database ... ok")

    # Load the table into a Python list
    devices = []

#    cursor.execute("SELECT id, CNF, CPU, RamAva, RamTot, DiskAva, DiskTot, IOPS, BW, Role FROM Provisioned_Devices")

    my_query = """
    SELECT 
        pd.Region, pd.Site, pd.ID AS "Device UUID", 
        pd.Status, pd.Role, pd.CNF,  
        pd.CPU AS "CPU Usage",
        pd.RamAva, pd.RamTot,
        pd.DiskAva, pd.DiskTot,
        pd.IOPS, pd.BW AS "Bandwidth", pd.Manufacturer, pd.Platform, 
        pd.Cluster, s.Location, s.ShortName, s.Type,
        s.Latitude,
        s.Longitude        
    FROM Provisioned_Devices pd 
    JOIN Sites s 
        ON pd.Site = s.ID 
    WHERE 
        pd.Region = 'DFW' 
        AND pd.Status = 'Active' 
        AND (pd.ID LIKE '%_UPF' OR pd.ID LIKE '%_DU') 
        AND (pd.Role = 'EdgeServer' OR pd.Role = 'FEServer');
    """
    cursor.execute(my_query)
    

    for row in cursor.fetchall():
        devices.append({
            "id": row[2], "CNF": row[5], "CPU": row[6], "RamAva": row[7], "RamTot": row[8],
            "DiskAva": row[9], "DiskTot": row[10], "IOPS": row[11], "BW": row[12], "Role": row[4],
            "cnf_value": None, "cpu_usage": None, "ramtot_value": None, "ramava_value": None,
            "diskava_value": None, "iops_value": None, "bw_value": None, "disktot_value": 4,
            "region": row[0], "site": row[1], "uuid": row[2], "status": row[3], "manufacturer": row[13],
            "platform": row[14], "cluster": row[15], "location": row[16], "shortname": row[17], 
            "type": row[18], "lat": row[19], "long": row[20],
            "distance": 0.00, "latency": 0.00
        })

    loguru.logger.debug(f"Loaded {len(devices)} devices from the database.")

    # First pass: Generate initial random values
    for device in devices:
        role = device["Role"]
        device["cnf_value"] = '1' if role == 'FEServer' else random.choice(['0', '1'])
        
        # CPU value -------------------------------------------------------------------------------
        if device["cnf_value"] == '1':
            device["cpu_usage"] = round(random.uniform(cpu_min_value_cnf_deployed, cpu_max_value_cnf_deployed), 2)
        else:
            device["cpu_usage"] = round(random.uniform(cpu_min_value_cnf_not_deployed, cpu_max_value_cnf_not_deployed), 2)

        
        # RAM total -------------------------------------------------------------------------------
        device["ramtot_value"] = 128 if role == 'FEServer' else random.choice([192, 256, 320])

        # Ram available ---------------------------------------------------------------------------
#        min_ramava = round(mem_min_thresold_value * device["ramtot_value"], 2)
#        max_ramava = round(mem_max_thresold_value * device["ramtot_value"], 2)

        if device["cnf_value"] == '1':
            min_ramava = round(mem_min_value_cnf_deployed * device["ramtot_value"], 2)
            max_ramava = round(mem_max_value_cnf_deployed * device["ramtot_value"], 2)
        else:
            min_ramava = round(mem_min_value_cnf_not_deployed * device["ramtot_value"], 2)
            max_ramava = round(mem_max_value_cnf_not_deployed * device["ramtot_value"], 2)
    
        device["ramava_value"] = round(random.uniform(min_ramava, max_ramava), 2)
       

        # Disk total ------------------------------------------------------------------------------
        device["disktot_value"] = disk_max_value

        # Disk available
#        min_diskava = round(disk_min_thresold_value * device["disktot_value"], 2)
#        max_diskava = round(disk_max_thresold_value * device["disktot_value"], 2)

        if device["cnf_value"] == '1':
            min_diskava = round(disk_min_value_cnf_deployed * device["disktot_value"], 2)  # 70% de 4 = 2.8
            max_diskava = round(disk_max_value_cnf_deployed * device["disktot_value"], 2)  # 85% de 4 = 3.4
        else:
            min_diskava = round(disk_min_value_cnf_not_deployed * device["disktot_value"], 2)  # 10% de 4 = 0.4
            max_diskava = round(disk_max_value_cnf_not_deployed * device["disktot_value"], 2)  # 40% de 4 = 1.6
        
#        device["diskava_value"] = round(random.uniform(disk_min_value, disk_max_value), 2)
        device["diskava_value"] = round(random.uniform(min_diskava, max_diskava), 2)


        # BW --------------------------------------------------------------------------------------
        device["bw_value"] = round(random.uniform(bw_min_value, bw_max_value),2) 


    loguru.logger.debug(f"First pass completed: Random values generated.")

    # Sort devices by CNF and CPU usage
#    devices.sort(key=lambda d: (d["Role"],d["cnf_value"], d["cpu_usage"], d["ramava_value"], d["diskava_value"]))

    # Generate unique decreasing IOPS values
    num_devices = len(devices)
    iops_values = sorted(random.sample(range(iops_min_value, iops_max_value), num_devices), reverse=True)

    for i, device in enumerate(devices):
        device["iops_value"] = iops_values[i]

    loguru.logger.debug(f"IOPS values assigned uniquely and in decreasing order.")

    # Ensure DiskAva does not exceed DiskTot
    for device in devices:
#        if device["diskava_value"] > 4:
#            device["diskava_value"] = 4.000
        if device["diskava_value"] > disk_max_value:
            device["diskava_value"] = disk_max_value2

    loguru.logger.debug(f"Disk availability values validated.")


    debug_modee = 1    
    if debug_modee == 1:
        headers = ["id", "cnf_value", "cpu_usage", "ramfree_value", "diskfree_value", "iops_value", "bw_value", "role", "distance", "latency"]

        table = []
        
        for device in [d for d in devices if '_CUUP' not in d['id'] or d['Role'] != 'CoreServer']:
            my_ramfree_value = round((device['ramava_value']/device['ramtot_value'])*100,2)
            my_diskfree_value = round((device['diskava_value']/device['disktot_value'])*100,2)

            current_device_long3 = device['long']
            current_device_lat3 = device['lat']

            device_location3 = (current_device_lat3, current_device_long3)
            
            # Coordinates of your position
            #my_location = (32.841362, -96.784582) # SMU Dallas Campus
            
            # Calculate distance: SMU to current FarEdge
            distance3 = round(haversine(my_location, device_location3, unit=Unit.MILES), 2)

            ### Latency ---------------------------------------------------------------------------
            ### 200,000 km/s (124,000 miles per second)
            ### latency (ms) = [ distance in miles ] / [ 124 ]
#            latency =  round((distance / 124), 4)
            #latency =  round((distance / 200), 4)
            #latency3 =  round((distance3 * 3.4), 2) # theoretical carrier latency per kilometer is about 3.4μs for radio and 5μs for fiber

            # f(lt/μs)=distance (mi)*7.9   -- fiber
            # f(lt/μs)=distance (mi)*5.37  -- microwave
            latency3 =  round((distance3 * 5.37), 2) 

            device['distance'] = distance3
            device['latency'] = latency3

            table.append([
                device['id'],
                device['cnf_value'],
                device['cpu_usage'],
                my_ramfree_value,
                my_diskfree_value,
                device['iops_value'],
                device['bw_value'],
                device['Role'],
                distance3,
                latency3
            ])


        debug_modee = 0    
        if debug_modee == 1:
            print(tabulate.tabulate(table, headers=headers, tablefmt="grid"))

        # Compute ramfree_value and diskfree_value
        for device in devices:
            device["ramfree_value"] = round((device["ramava_value"] / device["ramtot_value"]) * 100, 2)
            device["diskfree_value"] = round((device["diskava_value"] / device["disktot_value"]) * 100, 2)


        # Compute scoring
        score_data = []
        for device in [d for d in devices if d['Role'] == 'FEServer' and ('_CUUP' not in d['id'] ) ]:

            debug_modee = 0
            if debug_modee == 1:
                print(f"device_FEServer:{device}")
                
############cpu_score = sum(1 if device["cpu_usage"] < other["cpu_usage"] else -1 for other in devices if device != other)

            loguru.logger.debug(f"Calculating score for device {device['id']} with cpu_usage {device['cpu_usage']}")
            
            # Calculate the "corrected" value (comparison between all devices)
            comparisons = []
            for other in [d for d in devices if d['Role'] == 'FEServer' and ('_CUUP' not in d['id'] ) ]:
                if device != other:
                    comparisons.append(1 if device["cpu_usage"] < other["cpu_usage"] else -1)
            
            # Format the corrected sum as a string for display
            corrected = " + ".join([str(val) for val in comparisons])  # Sum format
            corrected_num = sum(comparisons)  # Actual sum to get the corrected value
            
            # Log the corrected sum and its result
            loguru.logger.debug(f"Device {device['id']} - Corrected: {corrected} = {corrected_num}")
            
            # Calculate the cpu_score by comparing each device's cpu_usage with others
            #cpu_score = sum(1 if device["cpu_usage"] < other["cpu_usage"] else -1 for other in devices if device != other)
#            cpu_score = corrected_num
            cpu_score = corrected_num * cpu_weight
            
            # Log the cpu_score result
            loguru.logger.debug(f"Device {device['id']} - cpu_score calculated: {cpu_score}")


#############ram_score = sum(1 if device["ramfree_value"] > other["ramfree_value"] else -1 for other in devices if device != other)
            loguru.logger.debug(f"Calculating score for device {device['id']} with ramfree_value {device['ramfree_value']}")
            
            # Calculate the "corrected" value (comparison between all devices)
            comparisons = []
            for other in [d for d in devices if d['Role'] == 'FEServer' and ('_CUUP' not in d['id'] ) ]:
                if device != other:
                    comparisons.append(1 if device["ramfree_value"] > other["ramfree_value"] else -1)
            
            # Format the corrected sum as a string for display
            corrected = " + ".join([str(val) for val in comparisons])  # Sum format
            corrected_num = sum(comparisons)  # Actual sum to get the corrected value
            
            # Log the corrected sum and its result
            loguru.logger.debug(f"Device {device['id']} - Corrected: {corrected} = {corrected_num}")
            
            # Calculate the ram_score by comparing each device's ramfree_value with others
            #ram_score = sum(1 if device["ramfree_value"] > other["ramfree_value"] else -1 for other in devices if device != other)
#            ram_score = corrected_num
            ram_score = corrected_num * ram_weight
            
            # Log the ram_score result
            loguru.logger.debug(f"Device {device['id']} - ramfree_value calculated: {ram_score}")

#############disk_score = sum(1 if device["diskfree_value"] > other["diskfree_value"] else -1 for other in devices if device != other)
            loguru.logger.debug(f"Calculating score for device {device['id']} with diskfree_value {device['diskfree_value']}")
            
            # Calculate the "corrected" value (comparison between all devices)
            comparisons = []
            for other in [d for d in devices if d['Role'] == 'FEServer' and ('_CUUP' not in d['id'] ) ]:
                if device != other:
                    comparisons.append(1 if device["diskfree_value"] > other["diskfree_value"] else -1)
            
            # Format the corrected sum as a string for display
            corrected = " + ".join([str(val) for val in comparisons])  # Sum format
            corrected_num = sum(comparisons)  # Actual sum to get the corrected value
            
            # Log the corrected sum and its result
            loguru.logger.debug(f"Device {device['id']} - Corrected: {corrected} = {corrected_num}")
            
            # Calculate the disk_score by comparing each device's diskfree_value with others
            #disk_score = sum(1 if device["diskfree_value"] > other["diskfree_value"] else -1 for other in devices if device != other)
#            disk_score = corrected_num
            disk_score = corrected_num * disk_weight
            
            # Log the disk_score result
            loguru.logger.debug(f"Device {device['id']} - diskfree_value calculated: {disk_score}")


#############iops_score = sum(1 if device["iops_value"] > other["iops_value"] else -1 for other in devices if device != other)
            loguru.logger.debug(f"Calculating score for device {device['id']} with iops_value {device['iops_value']}")
            
            # Calculate the "corrected" value (comparison between all devices)
            comparisons = []
            for other in [d for d in devices if d['Role'] == 'FEServer' and ('_CUUP' not in d['id'] ) ]:
                if device != other:
                    comparisons.append(1 if device["iops_value"] > other["iops_value"] else -1)
            
            # Format the corrected sum as a string for display
            corrected = " + ".join([str(val) for val in comparisons])  # Sum format
            corrected_num = sum(comparisons)  # Actual sum to get the corrected value
            
            # Log the corrected sum and its result
            loguru.logger.debug(f"Device {device['id']} - Corrected: {corrected} = {corrected_num}")
            
            # Calculate the iops_score by comparing each device's iops_value with others
            #iops_score = sum(1 if device["iops_value"] > other["iops_value"] else -1 for other in devices if device != other)
#            iops_score = corrected_num
            iops_score = corrected_num * iops_weight
            
            # Log the iops_score result
            loguru.logger.debug(f"Device {device['id']} - iops_value calculated: {iops_score}")

#############bw_score = sum(1 if device["bw_value"] > other["bw_value"] else -1 for other in devices if device != other)
            loguru.logger.debug(f"Calculating score for device {device['id']} with bw_value {device['bw_value']}")
            
            # Calculate the "corrected" value (comparison between all devices)
            comparisons = []
            for other in [d for d in devices if d['Role'] == 'FEServer' and ('_CUUP' not in d['id'] ) ]:
                if device != other:
                    comparisons.append(1 if device["bw_value"] > other["bw_value"] else -1)
            
            # Format the corrected sum as a string for display
            corrected = " + ".join([str(val) for val in comparisons])  # Sum format
            corrected_num = sum(comparisons)  # Actual sum to get the corrected value
            
            # Log the corrected sum and its result
            loguru.logger.debug(f"Device {device['id']} - Corrected: {corrected} = {corrected_num}")
            
            # Calculate the bw_score by comparing each device's bw_value with others
            #bw_score = sum(1 if device["bw_value"] > other["bw_value"] else -1 for other in devices if device != other)
#            bw_score = corrected_num
            bw_score = corrected_num * bw_weight
            
            # Log the bw_score result
            loguru.logger.debug(f"Device {device['id']} - bw_value calculated: {bw_score}")


#############latency_score = sum(1 if device["latency"] > other["latency"] else -1 for other in devices if device != other)
            loguru.logger.debug(f"Calculating score for device {device['id']} with latency {device['latency']}")
            
            # Calculate the "corrected" value (comparison between all devices)
            comparisons = []
            for other in [d for d in devices if d['Role'] == 'FEServer' and ('_CUUP' not in d['id'] ) ]:
                if device != other:
                    comparisons.append(1 if device["latency"] < other["latency"] else -1)
            
            # Format the corrected sum as a string for display
            corrected = " + ".join([str(val) for val in comparisons])  # Sum format
            corrected_num = sum(comparisons)  # Actual sum to get the corrected value
            
            # Log the corrected sum and its result
            loguru.logger.debug(f"Device {device['id']} - Corrected: {corrected} = {corrected_num}")
            
            # Calculate the latency_score by comparing each device's latency with others
            #latency_score = sum(1 if device["latency"] > other["latency"] else -1 for other in devices if device != other)
#            latency_score = corrected_num
            latency_score = corrected_num * latency_weight
            
            # Log the latency_score result
            loguru.logger.debug(f"Device {device['id']} - latency calculated: {latency_score}")


#            total_score = cpu_score + ram_score + disk_score + iops_score + bw_score  # Sum of all scores
            total_score = cpu_score + ram_score + disk_score + iops_score + bw_score + latency_score # Sum of all scores


            score_data.append({
                "region" : device['region'],
                "site" : device['site'],
                "id" : device['uuid'],
                "status" : device['status'],
                "role" : device['Role'],
                "cnf_value" : device['cnf_value'],
                "cpu_usage" : device['cpu_usage'],
                "cpu_score" : cpu_score,
                "ramfree_value" : device['ramfree_value'],
                "ram_score" : ram_score,
                "diskfree_value" : device['diskfree_value'],
                "disk_score" : disk_score,
                "iops_value" : device['iops_value'],
                "iops_score" : iops_score,
                "bw_value" : device['bw_value'],
                "bw_score" : bw_score,
                "latency_value" : device['latency'],
                "latency_score" : latency_score,
                "total_score" : total_score,
                "manufacturer" : device['manufacturer'],
                "platform" : device['platform'],
                "cluster" : device['cluster'],
                "shortname" : device['shortname'],
                "location" : device['location'],
                "type": device['type'],
                "long": device['long'],
                "lat": device['lat'],
                "distance" : device['distance']
            })

        # ____________________________________________________________________
        # Find the best FarEdge Server +++++++++++++++++++++++++++++++++++++++

        # Filter data by roles
        fe_server_data = [d for d in score_data if d['role'] == 'FEServer']

        best_server_faredge = get_best_edge_server(fe_server_data)

        # Take coordinates of closest FarEdge server
        latitude_fe2 = best_server_faredge[0]['lat']
        longitude_fe2 = best_server_faredge[0]['long']
    
        my_closest_fard_edge_is2 = (latitude_fe2,longitude_fe2)
        my_closest_fard_edge_name_is2 = best_server_faredge[0]['cluster']
        my_closest_fard_edge_name_is3 = best_server_faredge[0]['site']
        
        # ++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++

        for device in [d for d in devices if d['Role'] == 'EdgeServer' and ('_CUUP' not in d['id'] ) ]:

            debug_modee = 0
            if debug_modee == 1:
                print(f"device_EdgeServer:{device}")


            current_device_long = device['long']
            current_device_lat = device['lat']

            device_location2 = (current_device_lat, current_device_long)

            #utsw_coords  = (32.817195, -96.843869) ## UTSW Medical Center Dallas // W1225536259
            # Calculate distance from my current EdgeSite to UTSW Medical Center
            distance_current_edge_site_to_utsw = round(haversine(device_location2, utsw_coords, unit=Unit.MILES), 2)
            
            # Calculate distance from closes FarEdge site to my current EdgeSite
            distance2_tmp = round(haversine(my_closest_fard_edge_is2, device_location2, unit=Unit.MILES), 2)

            distance2 = round(distance_current_edge_site_to_utsw + distance2_tmp,2)
            debug_modee = 0
            if debug_modee == 1:
                loguru.logger.debug(f"(A) Distance from selected FarEdge to this EdgeSite: {distance2_tmp}miles")
                loguru.logger.debug(f"(B) Distance this EdgeSite to UTSW Medical Center  : {distance_current_edge_site_to_utsw}miles")
                loguru.logger.debug(f"Total distnace ( A + B )  : {distance2}miles")

            ### Latency ---------------------------------------------------------------------------
            ### 200,000 km/s (124,000 miles per second)
            ### latency (ms) = [ distance in miles ] / [ 124 ]
#            latency =  round((distance / 124), 4)
            #latency =  round((distance / 200), 4)
            #latency2 =  round((distance2 * 5), 2) # theoretical carrier latency per kilometer is about 3.4μs for radio and 5μs for fiber

            # f(lt/μs)=distance (mi)*7.9   -- fiber
            # f(lt/μs)=distance (mi)*5.37  -- microwave
            latency2 =  round((distance2 * 7.9), 2) 

            device['distance'] = distance2
            device['latency'] = latency2

                
############cpu_score = sum(1 if device["cpu_usage"] < other["cpu_usage"] else -1 for other in devices if device != other)

            loguru.logger.debug(f"Calculating score for device {device['id']} with cpu_usage {device['cpu_usage']}")
            
            # Calculate the "corrected" value (comparison between all devices)
            comparisons = []
            for other in [d for d in devices if d['Role'] == 'EdgeServer' and ('_CUUP' not in d['id'] ) ]:
                if device != other:
                    comparisons.append(1 if device["cpu_usage"] < other["cpu_usage"] else -1)
            
            # Format the corrected sum as a string for display
            corrected = " + ".join([str(val) for val in comparisons])  # Sum format
            corrected_num = sum(comparisons)  # Actual sum to get the corrected value
            
            # Log the corrected sum and its result
            loguru.logger.debug(f"Device {device['id']} - Corrected: {corrected} = {corrected_num}")
            
            # Calculate the cpu_score by comparing each device's cpu_usage with others
            #cpu_score = sum(1 if device["cpu_usage"] < other["cpu_usage"] else -1 for other in devices if device != other)
#            cpu_score = corrected_num
            cpu_score = corrected_num * cpu_weight
            
            # Log the cpu_score result
            loguru.logger.debug(f"Device {device['id']} - cpu_score calculated: {cpu_score}")


#############ram_score = sum(1 if device["ramfree_value"] > other["ramfree_value"] else -1 for other in devices if device != other)
            loguru.logger.debug(f"Calculating score for device {device['id']} with ramfree_value {device['ramfree_value']}")
            
            # Calculate the "corrected" value (comparison between all devices)
            comparisons = []
            for other in [d for d in devices if d['Role'] == 'EdgeServer' and ('_CUUP' not in d['id'] ) ]:
                if device != other:
                    comparisons.append(1 if device["ramfree_value"] > other["ramfree_value"] else -1)
            
            # Format the corrected sum as a string for display
            corrected = " + ".join([str(val) for val in comparisons])  # Sum format
            corrected_num = sum(comparisons)  # Actual sum to get the corrected value
            
            # Log the corrected sum and its result
            loguru.logger.debug(f"Device {device['id']} - Corrected: {corrected} = {corrected_num}")
            
            # Calculate the ram_score by comparing each device's ramfree_value with others
            #ram_score = sum(1 if device["ramfree_value"] > other["ramfree_value"] else -1 for other in devices if device != other)
#            ram_score = corrected_num
            ram_score = corrected_num * ram_weight
            
            # Log the ram_score result
            loguru.logger.debug(f"Device {device['id']} - ramfree_value calculated: {ram_score}")

#############disk_score = sum(1 if device["diskfree_value"] > other["diskfree_value"] else -1 for other in devices if device != other)
            loguru.logger.debug(f"Calculating score for device {device['id']} with diskfree_value {device['diskfree_value']}")
            
            # Calculate the "corrected" value (comparison between all devices)
            comparisons = []
            for other in [d for d in devices if d['Role'] == 'EdgeServer' and ('_CUUP' not in d['id'] ) ]:
                if device != other:
                    comparisons.append(1 if device["diskfree_value"] > other["diskfree_value"] else -1)
            
            # Format the corrected sum as a string for display
            corrected = " + ".join([str(val) for val in comparisons])  # Sum format
            corrected_num = sum(comparisons)  # Actual sum to get the corrected value
            
            # Log the corrected sum and its result
            loguru.logger.debug(f"Device {device['id']} - Corrected: {corrected} = {corrected_num}")
            
            # Calculate the disk_score by comparing each device's diskfree_value with others
            #disk_score = sum(1 if device["diskfree_value"] > other["diskfree_value"] else -1 for other in devices if device != other)
#            disk_score = corrected_num
            disk_score = corrected_num * disk_weight
            
            # Log the disk_score result
            loguru.logger.debug(f"Device {device['id']} - diskfree_value calculated: {disk_score}")


#############iops_score = sum(1 if device["iops_value"] > other["iops_value"] else -1 for other in devices if device != other)
            loguru.logger.debug(f"Calculating score for device {device['id']} with iops_value {device['iops_value']}")
            
            # Calculate the "corrected" value (comparison between all devices)
            comparisons = []
            for other in [d for d in devices if d['Role'] == 'EdgeServer' and ('_CUUP' not in d['id'] ) ]:
                if device != other:
                    comparisons.append(1 if device["iops_value"] > other["iops_value"] else -1)
            
            # Format the corrected sum as a string for display
            corrected = " + ".join([str(val) for val in comparisons])  # Sum format
            corrected_num = sum(comparisons)  # Actual sum to get the corrected value
            
            # Log the corrected sum and its result
            loguru.logger.debug(f"Device {device['id']} - Corrected: {corrected} = {corrected_num}")
            
            # Calculate the iops_score by comparing each device's iops_value with others
            #iops_score = sum(1 if device["iops_value"] > other["iops_value"] else -1 for other in devices if device != other)
#            iops_score = corrected_num
            iops_score = corrected_num * iops_weight
            
            # Log the iops_score result
            loguru.logger.debug(f"Device {device['id']} - iops_value calculated: {iops_score}")


#############bw_score = sum(1 if device["bw_value"] > other["bw_value"] else -1 for other in devices if device != other)
            loguru.logger.debug(f"Calculating score for device {device['id']} with bw_value {device['bw_value']}")
            
            # Calculate the "corrected" value (comparison between all devices)
            comparisons = []
            for other in [d for d in devices if d['Role'] == 'EdgeServer' and ('_CUUP' not in d['id'] ) ]:
                if device != other:
                    comparisons.append(1 if device["bw_value"] > other["bw_value"] else -1)
            
            # Format the corrected sum as a string for display
            corrected = " + ".join([str(val) for val in comparisons])  # Sum format
            corrected_num = sum(comparisons)  # Actual sum to get the corrected value
            
            # Log the corrected sum and its result
            loguru.logger.debug(f"Device {device['id']} - Corrected: {corrected} = {corrected_num}")
            
            # Calculate the bw_score by comparing each device's bw_value with others
            #bw_score = sum(1 if device["bw_value"] > other["bw_value"] else -1 for other in devices if device != other)
#            bw_score = corrected_num
            bw_score = corrected_num * bw_weight
            
            # Log the bw_score result
            loguru.logger.debug(f"Device {device['id']} - bw_value calculated: {bw_score}")


#############latency_score = sum(1 if device["latency"] > other["latency"] else -1 for other in devices if device != other)
            loguru.logger.debug(f"Calculating score for device {device['id']} with latency {device['latency']}")
            
            # Calculate the "corrected" value (comparison between all devices)
            comparisons = []
            for other in [d for d in devices if d['Role'] == 'EdgeServer' and ('_CUUP' not in d['id'] ) ]:
                if device != other:
                    comparisons.append(1 if device["latency"] < other["latency"] else -1)
            
            # Format the corrected sum as a string for display
            corrected = " + ".join([str(val) for val in comparisons])  # Sum format
            corrected_num = sum(comparisons)  # Actual sum to get the corrected value
            
            # Log the corrected sum and its result
            loguru.logger.debug(f"Device {device['id']} - Corrected: {corrected} = {corrected_num}")
            
            # Calculate the latency_score by comparing each device's latency with others
            #latency_score = sum(1 if device["latency"] > other["latency"] else -1 for other in devices if device != other)
#            latency_score = corrected_num
            latency_score = corrected_num * latency_weight
            
            # Log the latency_score result
            loguru.logger.debug(f"Device {device['id']} - latency calculated: {latency_score}")


#            total_score = cpu_score + ram_score + disk_score + iops_score + bw_score  # Sum of all scores
            total_score = cpu_score + ram_score + disk_score + iops_score + bw_score + latency_score # Sum of all scores


            score_data.append({
                "region" : device['region'],
                "site" : device['site'],
                "id" : device['uuid'],
                "status" : device['status'],
                "role" : device['Role'],
                "cnf_value" : device['cnf_value'],
                "cpu_usage" : device['cpu_usage'],
                "cpu_score" : cpu_score,
                "ramfree_value" : device['ramfree_value'],
                "ram_score" : ram_score,
                "diskfree_value" : device['diskfree_value'],
                "disk_score" : disk_score,
                "iops_value" : device['iops_value'],
                "iops_score" : iops_score,
                "bw_value" : device['bw_value'],
                "bw_score" : bw_score,
                "latency_value" : device['latency'],
                "latency_score" : latency_score,
                "total_score" : total_score,
                "manufacturer" : device['manufacturer'],
                "platform" : device['platform'],
                "cluster" : device['cluster'],
                "shortname" : device['shortname'],
                "location" : device['location'],
                "type": device['type'],
                "long": device['long'],
                "lat": device['lat'],
                "distance" : device['distance']
            })


        # Filter data by roles
        edge_server_data = [d for d in score_data if d['role'] == 'EdgeServer']
        
        headers = [
                "region",
                "site",
                "id",
                "status",
                "role",
                "cnf_value",
                "cpu_usage",
                "cpu_score",
                "ramfree_value",
                "ram_score",
                "diskfree_value",
                "disk_score",
                "iops_value",
                "iops_score",
                "bw_value",
                "bw_score",
                "latency_value",
                "latency_score",
                "total_score",
                "manufacturer",
                "platform",
                "cluster",
                "location",
                "shortname",
                "type",
                "long",
                "lat",
                "distance"
            ]

        debug_modee = 0
        if debug_modee == 1:
            # Show tables in console output ( for debug )
            print_table(edge_server_data, 'EdgeServer',headers)
            print_table(fe_server_data, 'FEServer',headers)

 
        debug_modee = 0
        if debug_modee == 1:
            print(edge_server_data)



        # ____________________________________________________________________
        # Find the best FarEdge Server +++++++++++++++++++++++++++++++++++++++
        # ++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
        best_server_faredge = get_best_edge_server(fe_server_data) 

        # Show best FarEdge Server
        debug_modee = 0    
        if debug_modee == 1:
            print("Best FarEdge Server:")
            print(best_server_faredge)

        debug_modee = 0
        if debug_modee == 1:
            # Show tables in console output ( for debug )
            print_table(best_server_faredge, 'BestFarEdgeServer',headers)



        # ____________________________________________________________________
        # Find the best Edge Server +++++++++++++++++++++++++++++++++++++++
        best_server_edge = get_best_edge_server(edge_server_data)
        # ++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++

        my_closest_edge_name_is2 = best_server_edge[0]['cluster']
        my_closest_edge_name_is3 = best_server_edge[0]['site']

        # Show best Edge Server
        debug_modee = 0    
        if debug_modee == 1:
            print(best_server_edge)


        debug_modee = 0
        if debug_modee == 1:
            # Show tables in console output ( for debug )
            print_table(best_server_edge, 'BestEdgeServer',headers)



        # ___________________________________________________________________
        # HTML---------------------------------------------------------------
        # +++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++

        # Function to generate tables with the headers and data
        def generate_table(role, data, color, my_case):
            html = ""
            html += "<table class='data-table'><tr>"
            
            match my_case: 
                case 0: # SMU          <-> FarEdge
                    label_for_column = "Latency (&microsec) | 5.37"
                case 1: # Best FarEdge <-> Edge
                    label_for_column = "Latency (&microsec) | 7.9"
                case 2: # Best FarEdge/Edge
                    label_for_column = "Latency (&microsec)"
                case 3: # Best FarEdge/Edge
                    label_for_column = "Latency (&microsec)"
                
    
            headers = [
                "Region",
                "Site",
                "Device UUID",
                "Status",
                "Role",
                "CNF",
                "CPU Usage %",
                "CPU Score",
                "RAM free %",
                "RAM Score",
                "Disk free %",
                "Disk Score",
                "IOPS",
                "IOPS Score",
                "BW (Gbps)",
                "BW Score",
                label_for_column,                
                "Latency Score",
                "TOTAL SCORE",
                "Manufacturer",
                "Platform",
                "Cluster",
                "Location",
                "ShortName",
                "Type",
                "Latitude",
                "Longitude",
                "Distance"
                ]


            row_class = ""

            edge_color = "#ffff99"  # Light yellow for EdgeServer
            fe_color = "#ffcc99"  # Light orange for FEServer
            
            html += "".join([f"<th>{header}</th>" for header in headers])
            html += "</tr>"

            total_score_style = "background-color: #ffe5cc; border: 1px solid black;"  # Naranja muy claro + borde suave

            
            # Adding rows with the device data
            for device in data:
                
                if role == 'FEServer': ## FarEdge
                    row_class = f"style='background-color: {fe_color};'" if device in (best_server_faredge) else ""
                elif role == 'EdgeServer': ## Edge
                    row_class = f"style='background-color: {edge_color};'" if device in (best_server_edge) else ""
                elif role == 'BestFarEdgeServer':
                    if device['role'] == 'FEServer':
                        row_class = f"style='background-color: {fe_color};'"# if device in (best_server_faredge) else ""
                    elif device['role'] == 'EdgeServer':
                        row_class = f"style='background-color: {edge_color};'" #if device in (best_server_edge) else ""
                    
                prefix_htlm = f"<tr {row_class}>"


                html += prefix_htlm
                html += f"<td>{device['region']}</td>"
                html += f"<td>{device['site']}</td>"
                html += f"<td>{device['id']}</td>"
                html += f"<td>{device['status']}</td>"
                html += f"<td>{device['role']}</td>"
                html += f"<td>{device['cnf_value']}</td>"
                html += f"<td>{device['cpu_usage']}</td>"
                html += f"<td>{device['cpu_score']}</td>"
                html += f"<td>{device['ramfree_value']}</td>"
                html += f"<td>{device['ram_score']}</td>"
                html += f"<td>{device['diskfree_value']}</td>"
                html += f"<td>{device['disk_score']}</td>"
                html += f"<td>{device['iops_value']:,}</td>"
                html += f"<td>{device['iops_score']}</td>"
                html += f"<td>{device['bw_value']}</td>"
                html += f"<td>{device['bw_score']}</td>"
                html += f"<td>{device['latency_value']}</td>"
                html += f"<td>{device['latency_score']}</td>"
                html += f"<td style='{total_score_style}'>{device['total_score']}</td>"
                html += f"<td>{device['manufacturer']}</td>"
                html += f"<td>{device['platform']}</td>"
                html += f"<td>{device['cluster']}</td>"
                html += f"<td>{device['location']}</td>"
                html += f"<td>{device['shortname']}</td>"
                html += f"<td>{device['type']}</td>"
                html += f"<td>{device['long']}</td>"
                html += f"<td>{device['lat']}</td>"
                html += f"<td>{device['distance']}</td>"
                html += "</tr>"
            
            html += "</table>"
            return html


        # ------------------------------------------------------------------------ #
        # Distance from 'SMU Dallas Campus' to 'Far Edge' sites:
        # ------------------------------------------------------------------------ #

        # Starting the HTML output with the header section
        all_data_html_1 = "<hr>"
        all_data_html_1 += f"""
        <div style="display: flex; justify-content: space-between; align-items: center;">
            <h2>Scoring __FarEdge__ sites ( from 'SMU Dallas Campus' to 'Far Edge' sites ): *</h2>
            <button onclick="window.scrollTo({{ top: 0, behavior: 'smooth' }})">Go to Top</button>
        </div>
        """

        # Generate HTML tables for both roles
        all_data_html_1 += "<hr>"
        all_data_html_1 += generate_table('FEServer', fe_server_data, 0, 0)

        debug_modee = 0    
        if debug_modee == 1:
            print(f"all_data_html_1:{all_data_html_1}")



        # ------------------------------------------------------------------------ #
        # Distance from '<closest FE>' FarEdge to EdgeDC sites:
        # ------------------------------------------------------------------------ #

        all_data_html_2 = "<hr>"
        fe_color = "#ffcc99"  # Light orange for FEServer
        
        all_data_html_2 += f"""
        <div style="display: flex; justify-content: space-between; align-items: center;">
            <h2>Scoring __Edge__ sites ( from '<span style="background-color: {fe_color}; color: black; padding: 2px 5px; border-radius: 4px;">
                {my_closest_fard_edge_name_is2}
            </span>' FarEdge to EdgeDC sites):**</h2>
            <button onclick="window.scrollTo({{ top: 0, behavior: 'smooth' }})">Go to Top</button>
        </div>
        """


        all_data_html_2 += generate_table('EdgeServer', edge_server_data, 0, 1)

        debug_modee = 0    
        if debug_modee == 1:
            print(f"all_data_html_2:{all_data_html_2}")
        


        # ------------------------------------------------------------------------ #
        # Summary Best FarEdge/Edge sites:
        # ------------------------------------------------------------------------ #

        # Starting the HTML output with the header section
        all_data_html_3 = "<hr>"
        all_data_html_3 += f"""
        <div style="display: flex; justify-content: space-between; align-items: center;">
            <h2>Summary best sites:***</h2>
            <button onclick="window.scrollTo({{ top: 0, behavior: 'smooth' }})">Go to Top</button>
        </div>
        """

        summary_best_servers = best_server_faredge + best_server_edge


        # Generate HTML tables for both roles
        all_data_html_3 += generate_table('BestFarEdgeServer', summary_best_servers,1,2)

        debug_modee = 0    
        if debug_modee == 1:
            print(f"all_data_html_3:{all_data_html_3}")


        # EXcel---------------------------------------------------------------

        headers = ["id", "cpu_usage", "cpu_score", "ramfree_value", "ram_score", "diskfree_value", "disk_score",
                   "iops_value", "iops_score", "bw_value", "bw_score", "latency_value", "latency_score", "total_score", "role"]

        # Filter data by roles
        edge_server_data = [d for d in score_data if d['role'] == 'EdgeServer']
        fe_server_data = [d for d in score_data if d['role'] == 'FEServer']

        # Convert the filtered data into pandas DataFrames
        edge_server_df = pd.DataFrame(edge_server_data, columns=headers)
        fe_server_df = pd.DataFrame(fe_server_data, columns=headers)

        # Generate a timestamp for the filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_folder = "output"  # Folder where the file will be saved

        # Ensure that the 'output' folder exists
        if not os.path.exists(output_folder):
            os.makedirs(output_folder)

        # Create the filename with the timestamp
        file_name = f"{output_folder}/Device_Scores_{timestamp}.xlsx"

        # Write both tables to the Excel file
        with pd.ExcelWriter(file_name) as writer:
            edge_server_df.to_excel(writer, sheet_name='EdgeServer', index=False)
            fe_server_df.to_excel(writer, sheet_name='FEServer', index=False)

        # Print the location of the saved Excel file
        debug_modee = 1
        if debug_modee == 1:
            loguru.logger.debug(f"The tables have been exported to: {file_name}")








    # Variable to control database update
    do_update = 1  # Set to 1 to perform update, 0 to skip

    if do_update:
        
        for device in devices:
            loguru.logger.debug(f"Updating device ID {device['id']}: CNF {device['CNF']} -> {device['cnf_value']}, CPU {device['CPU']} -> {device['cpu_usage']}, RamTot {device['RamTot']} -> {device['ramtot_value']}, RamAva {device['RamAva']} -> {device['ramava_value']}, DiskAva {device['DiskAva']} -> {device['diskava_value']}, IOPS {device['IOPS']} -> {device['iops_value']}, BW {device['BW']} -> {device['bw_value']}, DiskTot {device['DiskTot']} -> {device['disktot_value']}")
#            print(f"Device: {device}")
            cursor.execute(
                "UPDATE Provisioned_Devices SET CNF=?, CPU=?, RamTot=?, RamAva=?, DiskAva=?, IOPS=?, BW=?, DiskTot=? WHERE id=?",
                (device["cnf_value"], device["cpu_usage"], device["ramtot_value"], device["ramava_value"],
                 device["diskava_value"], device["iops_value"], device["bw_value"], device["disktot_value"], device["id"])
            )
                     
        conn.commit()
        
        loguru.logger.info("Database updated with the new values.")

    conn.close()

    return all_data_html_1, all_data_html_2, all_data_html_3, all_data_html_4, summary_best_servers, fe_server_data, edge_server_data



def save_html_to_file(html_content, filename=my_html_output2):
    """Save HTML to file."""
    with open(filename, "w", encoding="utf-8") as file:
        file.write(html_content)



## ++++++++++++++++++++++++++++++++++++++++++++++++++++++
## | Database Functions                                 | 
## ++++++++++++++++++++++++++++++++++++++++++++++++++++++
def execute_query_view(view_name: str):
    """Function to connect to the database and execute a query (view) """
    try:
        loguru.logger.debug(f"Execute view: 'SELECT * FROM {view_name}'")
        conn = sqlite3.connect(DATABASE_PATH)
        cursor = conn.cursor()
        cursor.execute(f"SELECT * FROM {view_name}")
        rows = cursor.fetchall()
        conn.close()
        return rows
    except sqlite3.OperationalError as e:
        raise HTTPException(status_code=400, detail=str(e))


def execute_sql_command(command: str):
    """Function to connect to the database and execute a query """
    try:
        debug_modee = 0
        if debug_modee == 1:
            loguru.logger.debug(f"Execute query: '{command}'")

        conn = sqlite3.connect(DATABASE_PATH)
        cursor = conn.cursor()
        cursor.execute(command)
        rows = cursor.fetchall()
        conn.commit()
        conn.close()

        return rows

    except sqlite3.OperationalError as e:
        raise HTTPException(status_code=400, detail=str(e))



## ++++++++++++++++++++++++++++++++++++++++++++++++++++++
## | Maps Functions                                     | 
## ++++++++++++++++++++++++++++++++++++++++++++++++++++++
def get_bezier_curve_points(
    start_lat_lon: tuple, end_lat_lon: tuple, control_point_position=0.5, control_point_offset=0.5, points=100) -> list:
    """Generate curve in map """

    # Function to calculate intermediate points along the Bezier curve
    def bezier_curve(p0, p1, p2, n_points=100):
        t = np.linspace(0, 1, n_points)
        points = np.zeros((len(t), 2))
        for i, _t in enumerate(t):
            points[i] = (1 - _t) ** 2 * p0 + 2 * (1 - _t) * _t * p1 + _t**2 * p2
        return points.tolist()

    start_lat_lon = np.array(start_lat_lon)
    end_lat_lon = np.array(end_lat_lon)

    # Calculate direction vector from End to Start
    direction_vector = end_lat_lon - start_lat_lon
    
    # Create control point that would be the pount where curliness occures
    control_point = start_lat_lon + direction_vector * control_point_position

    # Determine offset direction based on the offset factor
    if control_point_offset > 0:
        control_point += np.array([-direction_vector[1], direction_vector[0]]) * control_point_offset
    else:
        control_point += np.array([direction_vector[1], -direction_vector[0]]) * abs(control_point_offset)

    # Creating Bezier curve points
    curve_points = bezier_curve(start_lat_lon, control_point, end_lat_lon, points)
    return curve_points



## Prepare GUI
def render_gui(
    all_data_with_distances,
    closest_devices,
    my_closest_fard_edge_name_is,
    my_closest_edge_name_is,
    all_data_with_distances_edge,
    all_data_with_distances_all,
    all_data_html_1,
    all_data_html_2,
    all_data_html_3,
    all_data_html_4,
    my_summary_best_servers,
    my_fe_server_data,
    my_edge_server_data
):
    """Prepare GUI (HTML) """


    # List of buttons ( first row, top page )
    position_html = '<br>'
    position_html += """
        <button onclick="location.reload()">Refresh</button>
        <button onclick="document.getElementById('distance_fard_edge').scrollIntoView({ behavior: 'smooth' })">Distance SMU to FarEdge Info</button>
        <button onclick="document.getElementById('edge_dc_distance').scrollIntoView({ behavior: 'smooth' })">FE to EdgeDC Distance Info</button>
        <button onclick="document.getElementById('map_section_title').scrollIntoView({ behavior: 'smooth' })">Go to Map</button>
        <button onclick="document.body.style.zoom='67%'">Zoom 67%</button>
        <button onclick="document.body.style.zoom='100%'">Zoom 100%</button>

        <button id="exportButton">Excel</button>
        <div id="loading" style="
            display: none;
            position: fixed;
            top: 50%;
            left: 50%;
            transform: translate(-50%, -50%);
            font-size: 20px;
            font-weight: bold;
            background: rgba(0, 0, 0, 0.8);
            color: white;
            padding: 20px;
            border-radius: 10px;
            text-align: center;
            white-space: pre-line;
        ">
            ⏳  
            Generating Excel files...  
        </div>

        <button id="exportButton2">Excel & Download</button>
        <div id="loading2" style="
            display: none;
            position: fixed;
            top: 50%;
            left: 50%;
            transform: translate(-50%, -50%);
            font-size: 20px;
            font-weight: bold;
            background: rgba(0, 0, 0, 0.8);
            color: white;
            padding: 20px;
            border-radius: 10px;
            text-align: center;
            white-space: pre-line;
        ">
            ⏳  
            Generating Excel files and download...  
        </div>

        <div id="successMessage" style="
            min-width: 300px; /* Ancho mínimo */
            padding: 20px; /* Espaciado interno */
            text-align: center; /* Centrar el texto */
            display: none;
            position: fixed;
            top: 20px;
            left: 50%;
            transform: translateX(-50%);
            background-color: green;
            color: white;
            padding: 10px 20px;
            border-radius: 5px;
            font-size: 16px;
            z-index: 9999;
            opacity: 0;
            transition: opacity 0.5s ease;
        ">
            File generation OK
        </div>
        
        <div id="successMessage2" style="
            display: none;
            position: fixed;
            top: 20px;
            left: 50%;
            transform: translateX(-50%);
            background-color: green;
            color: white;
            padding: 10px 20px;
            border-radius: 5px;
            font-size: 16px;
            z-index: 9999;
            opacity: 0;
            transition: opacity 0.5s ease;
        ">
            File generation FAILED!
        </div>
        
        <script>
        
            document.getElementById("exportButton").addEventListener("click", async function() {
                let loadingDiv = document.getElementById("loading");
                loadingDiv.style.display = "block";  // Mostrar el mensaje de carga
            
                try {
                    let response = await fetch("/export_excel");
                    let data = await response.json();
                    let path = data.path; 
        
                    if (response.ok) {
            
                        let successMessage = document.getElementById("successMessage");
                        successMessage.innerText = `File generation OK:\n ${path}`;
                        successMessage.style.display = "block";
                        successMessage.style.opacity = 1;
            
                        // Ocultar el mensaje después de 2 segundos
                        setTimeout(() => {
                            successMessage.style.opacity = 0;
                            setTimeout(() => {
                                successMessage.style.display = "none";
                            }, 500);
                        }, 8000);
                        
                    } 
                    else {

                        // Mostrar el mensaje de éxito por 2 segundos
                        let successMessage2 = document.getElementById("successMessage2");
                        successMessage2.style.display = "block";
                        successMessage2.style.opacity = 1;  // Mostrar el mensaje con transición de opacidad
            
                        // Ocultar el mensaje después de 2 segundos
                        setTimeout(() => {
                            successMessage2.style.opacity = 0;
                            setTimeout(() => {
                                successMessage2.style.display = "none";  // Ocultar completamente después de la transición
                            }, 500); // Espera el tiempo de la transición
                        }, 2000);  // Espera 2 segundos antes de ocultarlo

                    
                    }
            
                } catch (error) {
                    alert("Server error: " + error);
                }
            
                loadingDiv.style.display = "none"; // Ocultar el mensaje de carga al finalizar
            });
            
            
            document.getElementById("exportButton2").addEventListener("click", async function() {
                let loadingDiv = document.getElementById("loading2");
                loadingDiv.style.display = "block";  // Mostrar el mensaje de carga
        
                try {
                    let response = await fetch("/export_excel_and_download");
                    let data = await response.json();
        
                    if (response.ok) {
                        let files = data.files;
                        
                        if (files.length === 0) {
                            alert("No files generated.");
                        } else {
                            files.forEach(file => {
                                let a = document.createElement("a");
                                a.href = file;
                                a.download = file.split('/').pop();  // Extraer el nombre del archivo
                                document.body.appendChild(a);
                                a.click();
                                document.body.removeChild(a);
                            });
                        }

                        // Mostrar el mensaje de éxito por 2 segundos
                        let successMessage = document.getElementById("successMessage");
                        successMessage.style.display = "block";
                        successMessage.style.opacity = 1;  // Mostrar el mensaje con transición de opacidad
            
                        // Ocultar el mensaje después de 2 segundos
                        setTimeout(() => {
                            successMessage.style.opacity = 0;
                            setTimeout(() => {
                                successMessage.style.display = "none";  // Ocultar completamente después de la transición
                            }, 500); // Espera el tiempo de la transición
                        }, 2000);  // Espera 2 segundos antes de ocultarlo


                    } else {
                        //alert("Error generating Excel files.");

                        // Mostrar el mensaje de éxito por 2 segundos
                        let successMessage2 = document.getElementById("successMessage2");
                        successMessage2.style.display = "block";
                        successMessage2.style.opacity = 1;  // Mostrar el mensaje con transición de opacidad
            
                        // Ocultar el mensaje después de 2 segundos
                        setTimeout(() => {
                            successMessage2.style.opacity = 0;
                            setTimeout(() => {
                                successMessage2.style.display = "none";  // Ocultar completamente después de la transición
                            }, 500); // Espera el tiempo de la transición
                        }, 2000);  // Espera 2 segundos antes de ocultarlo


                    }
                } catch (error) {
                    alert("Server error: " + error);
                }
        
                loadingDiv.style.display = "none"; // Ocultar el mensaje de carga al finalizar
            });

        </script>
        
    """
        
    # ------------------------------------------------------------------------ #
    # Coordinates of 'SMU Dallas Campus'
    position_html += '<hr>'
    position_html += f"""
    <div style="display: flex; justify-content: space-between; align-items: center;">
        <h2>Coordinates of 'SMU Dallas Campus':</h2>
        <button onclick="window.scrollTo({{ top: 0, behavior: 'smooth' }})">Go to Top</button>
    </div>
    <p>Latitude: {my_location[0]}, Longitude: {my_location[1]}</p>
    """

    # ------------------------------------------------------------------------ #
    # All Data from '<view_name>':
    # All Data from 'UOP_1_Shared_UPF_Edge_Far_Edge_DCs_location_info_v3':
    # ------------------------------------------------------------------------ #
    all_data_html = "<hr>"
    all_data_html += f"""
    <div style="display: flex; justify-content: space-between; align-items: center;">
        <h2>All Data from '{view_name}':</h2>
        <button onclick="window.scrollTo({{ top: 0, behavior: 'smooth' }})">Go to Top</button>
    </div>
    <table class='data-table'><tr>
    """

    # Headers for table 'All Data from 'view_name':' 
    headers = [
        "Region",
        "Site",
        "Device UUID",
        "Status",
        "Role",
        "CNF",
        "CPU Usage %",
        "RAM free %",
        "Disk free %",
        "IOPS",
        "Bandwidth (Gbps)",
        "Manufacturer",
        "Platform",
        "Cluster",
        "Location",
        "ShortName",
        "Type",
        "Latitude",
        "Longitude"
    ]
    all_data_html += "".join([f"<th>{header}</th>" for header in headers])
    all_data_html += "</tr>"
    
    # Track the colors for highlighting
    edge_color = "#ffff99"  # Light yellow for EdgeServer
    fe_color = "#ffcc99"  # Light orange for FEServer

    # Clean up contents for 'Distance_SMU_to_far_edge_table' table
    my_command = f"DELETE FROM Distance_SMU_to_far_edge_table;"
    data = execute_sql_command(my_command)

    debug_modee = 0
    if debug_modee == 1:
        print(f"{all_data_with_distances=}")

    all_tuples = []

    for row in my_fe_server_data:
        my_fe_server_data_tuple = (
            row.get('total_score'),
            row.get('region'),
            row.get('site'),
            row.get('id'),
            row.get('status'),
            row.get('role'),
            row.get('cnf_value'),
            row.get('cpu_usage'),
            row.get('ramfree_value'),
            row.get('diskfree_value'),
            row.get('iops_value'),
            row.get('bw_value'),
            row.get('manufacturer'),
            row.get('platform'),
            row.get('cluster'),
            row.get('location'),       
            row.get('shortname'),      
            row.get('type'),           
            row.get('lat'),
            row.get('long'),
            row.get('distance'),
            row.get('latency_value'),
            row.get('cpu_score'),
            row.get('ram_score'),
            row.get('disk_score'),
            row.get('iops_score'),
            row.get('bw_score'),
            row.get('latency_score'),
            row.get('total_score')     
        )
        all_tuples.append(my_fe_server_data_tuple)
        
        
    for row in my_edge_server_data:
        my_edge_server_data_tuple = (
            row.get('total_score'),
            row.get('region'),
            row.get('site'),
            row.get('id'),
            row.get('status'),
            row.get('role'),
            row.get('cnf_value'),
            row.get('cpu_usage'),
            row.get('ramfree_value'),
            row.get('diskfree_value'),
            row.get('iops_value'),
            row.get('bw_value'),
            row.get('manufacturer'),
            row.get('platform'),
            row.get('cluster'),
            row.get('location'),       
            row.get('shortname'),      
            row.get('type'),           
            row.get('lat'),
            row.get('long'),
            row.get('distance'),
            row.get('latency_value'),
            row.get('cpu_score'),
            row.get('ram_score'),
            row.get('disk_score'),
            row.get('iops_score'),
            row.get('bw_score'),
            row.get('latency_score'),
            row.get('total_score')     
        )

        all_tuples.append(my_edge_server_data_tuple)
    

    # Convertir lista a tupla si realmente necesitas que sea tipo `tuple`
    final_combined_tuple = tuple(all_tuples)



    # Insert values in 'Distance_SMU_to_far_edge_table' table
    for row in final_combined_tuple:
        
        sql_query = f"""
        INSERT INTO Distance_SMU_to_far_edge_table
            (
            Score, Region, Site, Device_UUID, Status, Role, CNF, CPU, 
            RAM_free, Disk_free, IOPS, 
            BW, Manufacturer, Platform, Cluster, 
            Location, ShortName, Type, Latitude, Longitude,
            Distance, Latency, 
            cpu_score, mem_score, disk_score, iops_score, bw_score, latency_score, final_score
            )    
        VALUES
        {row};
        """

        data = execute_sql_command(sql_query)


    
    # Define a sorting key function: FEServer comes first (0), then EdgeServer (1), and then sort by site name
    def server_type_sort_key(row):
        """Order by role """
        return (0 if row[4] == 'FEServer' else 1, row[1])

    # Sort the list using the custom key
    sorted_data = sorted(all_data_with_distances_all, key=server_type_sort_key)


    # Create HTML table
    for row in sorted_data:

        # Determine row class based on closest device type
        row_class = ""
        if row == closest_devices.get("EdgeServer"):
            row_class = f"style='background-color: {edge_color};'"
        elif row == closest_devices.get("FEServer"):
            row_class = f"style='background-color: {fe_color};'"
        
        # Convert tuple in list 
        row_list = list(row)

        # Remove last element (not needed)
        row_list.pop()

        # Convert back list into tuple 
        row = tuple(row_list)

        # Get the short_name for the onclick
        short_name = row_list[15]
                
        all_data_html += f"<tr {row_class}>" + "".join(
            [f"<td>{format(int(item), ',') if i == 9 and isinstance(item, (int, float)) else item}</td>" for i, item in enumerate(row_list)]
        ) + "</tr>"

    
    
    all_data_html += "</table>"

    debug_modee = 0    
    if debug_modee == 1:
        print(f"all_data_html_1:{all_data_html_1}")

    debug_modee = 0    
    if debug_modee == 1:
        print(f"all_data_html_2:{all_data_html_2}")

    debug_modee = 0    
    if debug_modee == 1:
        print(f"all_data_html_3:{all_data_html_3}")

    debug_modee = 0    
    if debug_modee == 1:
        print(f"all_data_html_4:{all_data_html_4}")


    all_data_html += all_data_html_1
    all_data_html += all_data_html_2
    all_data_html += all_data_html_3
    all_data_html += all_data_html_4
    


    # ------------------------------------------------------------------------------------------
    # Create the map with folium
    # ------------------------------------------------------------------------------------------
    map_path = "html/map.html"
    map_html = create_map(all_data_with_distances, closest_devices, all_data_with_distances_edge, my_summary_best_servers, sorted_data)
    #print(map_html)
    # Save the HTML content to a file
    save_html_to_file(map_html, filename=map_path)


    html_content = f"""
    <html>
    <head>
        <title>UOP v1.0</title>    
        <style>
            body {{
                font-family: Arial, sans-serif;
                margin: 20px;
            }}
            table {{
                width: 100%;
                border-collapse: collapse;
                margin-top: 20px;
            }}
            .data-table th, .data-table td {{
                border: 1px solid #ddd;
                padding: 8px;
                text-align: left;
            }}
            .data-table th {{
                background-color: #f4f4f4;
                font-weight: bold;
                white-space: nowrap;
            }}
            .data-table td {{
                white-space: nowrap;
                overflow: hidden;
                text-overflow: ellipsis;
            }}
            h2 {{
                font-size: 18px;
                margin-top: 20px;
            }}
            table {{
                font-size: 14px;
            }}
            #map {{
                width: 100%;
                height: 300px;  /* Adjusted height */
                margin-top: 20px;
            }}
        </style>

        <script>
            function setZoom(zoomLevel) {{
                document.body.style.zoom = zoomLevel;
            }}
        </script>

    </head>

    <body>
        {position_html}
        {all_data_html}
        <div style="display: flex; justify-content: space-between; align-items: center;">
            <h2 id="map_section_title">Map:</h2>
            <button onclick="window.scrollTo({{ top: 0, behavior: 'smooth' }})">Go to Top</button>
        </div>
        <div id="map" style="height: 300px; width: 100%;">{map_html}</div>
    </body>

    </html>
    """

    
    # Save the HTML content to a file
    save_html_to_file(html_content, filename=my_html_output2)

    return html_content



def create_map(all_data_with_distances, closest_devices,all_data_with_distances_edge,my_summary_best_servers,my_sorted_data):
    """Create Map """

    # Create a map centered around the new center location
    m = folium.Map(location=my_location, zoom_start=13)
    
    
    # Add markers for all devices
    marker_cluster = MarkerCluster().add_to(m)

    for row in my_sorted_data:

        latitude_value = float(row[17])  # Correct column for Latitude
        longitude_value = float(row[18])  # Correct column for Longitude
        location_value = row[14]  # Column with Location information
        short_name_value = row[15]  # Column with ShortName information
        type_server_value = row[16] # Column with Type_2: EDGE or FEDGE

        latitude = latitude_value
        longitude = longitude_value
        location = location_value
        short_name = short_name_value
        type_server = type_server_value
        
        if type_server == 'EDGE':
#                my_marker_color = 'red' 
            my_marker_color = 'purple' 
            my_marker_icon = 'cloud' 
        else: # FEDGE
#                my_marker_color = '#f5cc87' # 'orange'
            my_marker_color = 'orange'
            my_marker_icon = 'signal' 


        # Adding sites to map
        if type_server == 'EDGE':

            folium.Marker(
                location=(latitude, longitude),
                tooltip=f"{location} ({short_name})",
                icon=folium.Icon(icon=my_marker_icon, prefix='fa', color=my_marker_color)  # Red color for other markers
            ).add_to(marker_cluster)


        else:
        
            my_icon_file = 'images/logo.png'
            folium.Marker(
                location=(latitude, longitude),
                tooltip=f"{location} ({short_name})",
                #icon=folium.Icon(icon=my_marker_icon, prefix='fa', color=my_marker_color)  # Red color for other markers
                icon=folium.features.CustomIcon(icon_image=my_icon_file, icon_size=(29, 37)),  # Usa ícono personalizado con nuevo tamaño
            ).add_to(marker_cluster)
#            marker = folium.Marker(
#                location=(latitude, longitude),
#                tooltip=f"{location} ({short_name})",
#                #icon=folium.Icon(icon=my_marker_icon, prefix='fa', color=my_marker_color)  # Red color for other markers
#                icon=folium.features.CustomIcon(icon_image=my_icon_file, icon_size=(29, 37)),  # Usa ícono personalizado con nuevo tamaño
#            )


    # Add a marker for 'SMU Dallas Campus' location
    folium.Marker(
        location=my_location,
        tooltip="SMU Dallas Campus",
        icon=folium.Icon(icon='mortar-board', prefix='fa', color='blue')
    ).add_to(m)

    edge_coords = (my_summary_best_servers[1]['lat'],my_summary_best_servers[1]['long'])
    edge_distance = my_summary_best_servers[1]['distance']
    far_edge_coords = (my_summary_best_servers[0]['lat'],my_summary_best_servers[0]['long'])
    far_edge_distance = my_summary_best_servers[0]['distance']

    smu_coords = (32.841362, -96.784582)  # SMU Dallas Campus
    far_edge_coords = (my_summary_best_servers[0]['lat'], my_summary_best_servers[0]['long'])  # Far Edge Server
    edge_coords = (my_summary_best_servers[1]['lat'], my_summary_best_servers[1]['long'])  # Edge Server
    utsw_coords = (32.817195, -96.843869)  # UTSW Medical Center

    def mas_al_norte(coord1, coord2):
        if coord1[0] > coord2[0]:
            control_point_offset = 0.5
        elif coord1[0] < coord2[0]:
            control_point_offset = -0.5
        else:
            control_point_offset = 0.5
        return control_point_offset

    ####################################################################
    ## 
    ## SMU Campus < -- > Closest Far Edge 
    ## 
    ####################################################################


    # Añadir un círculo en el mapa
    folium.Circle(
        location=far_edge_coords,  # Coordenadas del centro del círculo
        radius=400,            # Radio del círculo en metros
        color='blue',          # Color del borde del círculo
        fill=True,             # Rellenar el círculo
        fill_color='blue',     # Color de relleno
        fill_opacity=0.2       # Opacidad del relleno
    ).add_to(m)


    smu_coords = my_location    
    my_control_point_offset = mas_al_norte(smu_coords,far_edge_coords)
    # Distances (curves) -----------------------------------------------------------------
    curve_points = get_bezier_curve_points(
        smu_coords,  # SMUD Campus
        far_edge_coords, # Far Edge
#        control_point_offset=-0.5
#        control_point_offset=0.5
        control_point_offset=my_control_point_offset
    )

    # Adding Bezier curve to the map
    curve_line = folium.PolyLine(
#        weight=55,
        curve_points[::-1], 
#        color="purple", 
#        color="green", 
        color="blue", 
        weight=10,
        opacity=0.4,
        tooltip='Distance: ' + str(far_edge_distance) +'km'
        ).add_to(m)


    add_plane = 0
    if add_plane == 1:
        # Add plane to the line
    #    attr = {"fill": "purple", "font-weight": "bold", "font-size": "30"}
        attr = {"fill": "green", "font-weight": "bold", "font-size": "30"}
        plugins.PolyLineTextPath(
            curve_line,
    ##        "\u2708",  # Plane unicode symbol
    ##        "\u25BA",  # Small arrow unicode symbol
            "\u2192",  # Normal arrow unicode symbol
            repeat=False,
            offset=14.5,
            orientation=180,
            attributes=attr,
        ).add_to(m)


    ####################################################################
    ## 
    ## Closest Far Edge < -- >  Closest Edge
    ## 
    ####################################################################

    # Añadir un círculo en el mapa
    folium.Circle(
        location=edge_coords,  # Coordenadas del centro del círculo
        radius=400,            # Radio del círculo en metros
        color='red',          # Color del borde del círculo
        fill=True,             # Rellenar el círculo
        fill_color='red',     # Color de relleno
        fill_opacity=0.2       # Opacidad del relleno
    ).add_to(m)


    # Distances (curves) -----------------------------------------------------------------
    my_control_point_offset = mas_al_norte(far_edge_coords,edge_coords)
    curve_points = get_bezier_curve_points(
        far_edge_coords, # Far Edge
        edge_coords,  # Edge
#        control_point_position=0.9, 
#        control_point_offset=0.5
#        control_point_offset=-0.5
        control_point_offset=my_control_point_offset
    )
    # Adding Bezier curve to the map
    curve_line = folium.PolyLine(
        curve_points[::-1], 
#        color="purple", 
#        color="green", 
        color="red", 
        weight=10,
        opacity=0.4,
        tooltip='Distance: ' + str(edge_distance) +'km'
        ).add_to(m)


    add_plane = 0
    if add_plane == 1:
        # Add plane to the line
    #    attr = {"fill": "purple", "font-weight": "bold", "font-size": "30"}
        attr = {"fill": "green", "font-weight": "bold", "font-size": "30"}
        plugins.PolyLineTextPath(
            curve_line,
    ##        "\u2708",  # Plane unicode symbol
    ##        "\u25BA",  # Small arrow unicode symbol
            "\u2192",  # Normal arrow unicode symbol
            repeat=False,
            offset=14.5,
            orientation=180,
            attributes=attr,
        ).add_to(m)


    # Calculate distance
    utsw_coords  = (32.817195, -96.843869) ## UTSW Medical Center Dallas // W1225536259 
    distance_edge_to_utsw = round(haversine(edge_coords,utsw_coords, unit=Unit.MILES), 2)
    
    my_arli_info = 'UTSW Medical Center Dallas'
    folium.Marker(
        location=utsw_coords,
        icon=folium.Icon(icon='medkit', prefix='fa', color='green'), 
#        popup=name,
        tooltip=my_arli_info
    ).add_to(m)





    ####################################################################
    ## 
    ## Closest Edge < -- >  UTSW Medical Center Dallas
    ## 
    ####################################################################
    # Distances (curves) -----------------------------------------------------------------
    my_control_point_offset = mas_al_norte(edge_coords,utsw_coords)
    curve_points = get_bezier_curve_points(
        edge_coords,  # Edge
        utsw_coords, # UTSW Medical Center Dallas
#        control_point_position=0.9, 
#        control_point_offset=0.2
#        control_point_offset=0.5
         control_point_offset=my_control_point_offset
   )
    # Adding Bezier curve to the map
    curve_line = folium.PolyLine(
        curve_points[::-1], 
#        color="purple", 
        color="green", 
        weight=10,
        opacity=0.4,
            
        tooltip='Distance: ' + str(distance_edge_to_utsw) +'km'
        ).add_to(m)

    add_plane = 0
    if add_plane == 1:
        # Add plane to the line
    #    attr = {"fill": "purple", "font-weight": "bold", "font-size": "30"}
        attr = {"fill": "green", "font-weight": "bold", "font-size": "30"}
        plugins.PolyLineTextPath(
            curve_line,
            "\u2708",  # Plane unicode symbol
            repeat=False,
            offset=14.5,
            orientation=180,
            attributes=attr,
        ).add_to(m)


    utsw_coords  = (32.817195, -96.843869) ## UTSW Medical Center Dallas // W1225536259 
    # Calculate distance
#    distance_utsw_arli = round(haversine(utsw_coords, arlington_coords, unit=Unit.MILES), 2)


    my_arli_info = 'Arlington National DC'
    folium.Marker(
        location=arlington_coords,
        icon=folium.Icon(icon='cubes', prefix='fa', color='darkblue'), 
#        popup=name,
        tooltip=my_arli_info
    ).add_to(m)
    
    my_arli_info = 'UTSW Medical Center Dallas'
    folium.Marker(
        location=utsw_coords,
        icon=folium.Icon(icon='medkit', prefix='fa', color='green'), 
#        popup=name,
        tooltip=my_arli_info
    ).add_to(m)



    # Generate the HTML for the map
    return m._repr_html_()

        
        
## ++++++++++++++++++++++++++++++++++++++++++++++++++++++
## | Excel Functions                                    | 
## ++++++++++++++++++++++++++++++++++++++++++++++++++++++
def set_first_row_gray(excel_file):
    """Sets the background of the first row to gray (#F2F2F2) only for columns that contain data."""
    try:
        # Load the existing Excel file
        wb = load_workbook(excel_file)
        
        # Get the active sheet
        ws = wb.active
        
        # Define the gray fill color
        fill_gray = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        # Get the number of columns with data in the first row
        max_col = ws.max_column
        
        for col in range(2, max_col + 1):
            # If the cell has data, set the background to gray
            if ws.cell(row=2, column=col).value is not None:
                ws.cell(row=2, column=col).fill = fill_gray
        
        # Save the modified Excel file
        wb.save(excel_file)
        
        debug_modee = 0
        if debug_modee == 1:
            print(f"First row background color set to gray for columns with data in {excel_file}")
        
    except Exception as e:
        print(f"Error: {e}")



def set_background_white(excel_file):
    """Opens an Excel file and sets the background of all cells in the active sheet to white."""
    try:
        # Load the existing Excel file
        wb = load_workbook(excel_file)
        
        # Get the active sheet
        ws = wb.active
        
        # Define the white fill color
        fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        # Set the maximum row and column as a large range to cover the entire sheet
        max_row = 1000  # You can adjust this to a higher value if needed
        max_col = 50  # You can adjust this to a higher value if needed
        
        # Apply white background to all cells in the sheet, covering a large range
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.fill = fill_white
        
        # Save the modified Excel file
        wb.save(excel_file)

        debug_modee = 0
        if debug_modee == 1:
            print(f"Background color set to white for all cells in {excel_file}")
        
        set_first_row_gray(excel_file)
        
    except Exception as e:
        print(f"Error: {e}")



def apply_number_format(ws, col, number_format):
    """Applies a number format to a specific column."""
    for row in range(2, ws.max_row + 1):  # Start from row 2 to avoid the header
        cell = ws.cell(row=row, column=col)
        cell.number_format = number_format



# Helper function to highlight the highest scores for FEServer and EdgeServer
def highlight_high_scores(excel_file):
    """Highlights the rows with the highest scores for FEServer and EdgeServer"""

    global my_closest_fard_edge_name_is2
    global my_closest_edge_name_is2
    global my_closest_fard_edge_name_is3
    global my_closest_edge_name_is3

    try:
        # Load the Excel file
        wb = load_workbook(excel_file)
        ws = wb.active
        
        # Define the highlight colors
        yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")  # Light Yellow
        orange_fill = PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid")  # Light Orange
        
        # Variables to track the highest scores and their corresponding rows
        feserver_row = None
        edgeserver_row = None
        
        # Start from row 3 (index 3 in 1-based indexing, row 2 in 0-based)
        for row in range(3, ws.max_row + 1):
            role = ws.cell(row=row, column=5).value  # Column E (Role)

            site = ws.cell(row=row, column=3).value  # Column C (Site)

            if role == 'FEServer' and site == my_closest_fard_edge_name_is3 :
                feserver_row = row
            elif role == 'EdgeServer'  and site == my_closest_edge_name_is3:
                edgeserver_row = row
            
        
        # Highlight the rows with the highest scores for FEServer and EdgeServer
        if feserver_row:
            for cell in ws[feserver_row]:
                cell.fill = orange_fill
        if edgeserver_row:
            for cell in ws[edgeserver_row]:
                cell.fill = yellow_fill
        
        # Save the modified Excel file
        highlighted_file = excel_file.replace('.xlsx', '_highlighted.xlsx')
        wb.save(highlighted_file)

        debug_modee = 0
        if debug_modee == 1:
            loguru.logger.debug(f"File saved as: {highlighted_file}")
    
    except Exception as e:
        print(f"Error: {e}")

        

def export_view_to_excel(view_name):
    """Executes the view and exports the results to an Excel file"""
    try:
        # Connect to the database and execute the query
        conn = sqlite3.connect(DATABASE_PATH)
        df = pd.read_sql_query(f"SELECT * FROM {view_name}", conn)
        conn.close()
        
        # Create an Excel file
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_path = os.path.join(output_excel, f"{view_name}_{timestamp}.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_names.get(view_name, "Sheet1")

        # Insert a row at the top (row 1) and set its height
        ws.insert_rows(1)  # Insert one row at the top
        ws.row_dimensions[1].height = 18  # Set row height for the first row
        
        # Insert a column at the left (column A) and set its width
        ws.insert_cols(1)  # Insert one column at the left
        ws.column_dimensions['A'].width = 8.43  # Set column width for the first column

        # Set background color to white
        fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        fill_gray = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        # Write data to the sheet starting from B2 (second row and second column)
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 2):
            # Write each row starting from the second row (r_idx starts at 2)
            for c_idx, cell_value in enumerate(row, 2):  # c_idx starts at 2 to begin writing in column B
                ws.cell(row=r_idx, column=c_idx, value=cell_value)
            ws.row_dimensions[r_idx].height = 18  # Set row height
            
            # Set header background color (only for the first data row, which corresponds to row 2 in Excel)
            if r_idx == 2:  # The first row of data (which corresponds to the second row in Excel)
                for cell in ws[r_idx]:
                    cell.fill = fill_gray
        
        # Format cells
        font = Font(name="Times New Roman", size=9)
        alignment = Alignment(horizontal="center", vertical="center")
        border = Border(left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin'))
        
        # Apply styles to all cells starting from B2 (i.e., from the second row and second column)
        max_col = ws.max_column
        max_row = ws.max_row
        
        for row in range(2, max_row + 1):  # Start from row 2
            for col in range(2, max_col + 1):  # Start from column 2 (B)
                cell = ws.cell(row=row, column=col)
                cell.font = font
                cell.alignment = alignment
                cell.border = border
                cell.fill = fill_white  # Apply white background to all cells

        # Apply number formatting based on the table names and columns
        if view_name == "Table_II_UOP_VIEW_0_Nearby_Cell_Sites_Far_Edge_DCs" or \
           view_name == "Table_V_UOP_VIEW_3_WITH_EDCs_FOR_A_DEDICATED_UPF":
            # Columns I: Thousand separator
            apply_number_format(ws, 9, '#,##0')  # Column I is 9th column
            # Columns F, G, H, J: Decimal with 2 places
            for col in [6, 7, 8, 10]:  # Columns F, G, H, J
                apply_number_format(ws, col, '#,##0.00')
        
        elif view_name == "Table_III_UOP_VIEW_1_WITH_EDCs_FOR_A_SHARED_UPF":
            # Columns I: Thousand separator
            apply_number_format(ws, 9, '#,##0')  # Column I is 9th column
            # Columns F, G, H, J: Decimal with 2 places
            for col in [6, 7, 8, 10]:  # Columns F, G, H, J
                apply_number_format(ws, col, '#,##0.00')
        
        elif view_name == "Table_IV_UOP_VIEW_2_WITH_EDCs_FOR_A_SHARED_UPF_with_NSSAI":
            # Column K: Thousand separator
            apply_number_format(ws, 11, '#,##0')  # Column K is 11th column
            # Columns H, I, J, L: Decimal with 2 places
            for col in [8, 9, 10, 12]:  # Columns H, I, J, L
                apply_number_format(ws, col, '#,##0.00')

        elif view_name == "Table_VI_UOP_VIEW_4A_WITH_FE_EDCs_FOR_A_SHARED_UPF":
            # Column J: Thousand separator
            apply_number_format(ws, 10, '#,##0')  # Column J is 10th column
            # Columns G, H, I, K: Decimal with 2 places
            for col in [7, 8, 9, 11]:  # Columns G, H, I, K
                apply_number_format(ws, col, '#,##0.00')
        
        # Auto-adjust column width for columns B and onwards
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 2  # Auto-adjust column width
        
        # Save the Excel file
        wb.save(file_path)
        
        debug_modee = 0
        if debug_modee == 1:
            loguru.logger.debug(f"File generated: {file_path}")
        
        set_background_white(file_path)
        # Apply background color changes

        if view_name == "Table_VI_UOP_VIEW_4A_WITH_FE_EDCs_FOR_A_SHARED_UPF":
            # Highlight the highest scores for FEServer and EdgeServer
            highlight_high_scores(file_path)

        return file_path
        
    except Exception as e:
        print(f"Error exporting {view_name}: {e}")        





####################################################################
## 
## FastAPI methods 
## 
####################################################################

# +++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
@app.get("/test")
async def get_hello_world():
    """Hello world"""
    loguru.logger.info("Received: GET /test")

    # Log de nivel INFO
    loguru.logger.info("Hello world! Sample info log")

    # Log de nivel DEBUG
    loguru.logger.debug("Hello world! Sample debug log")

    return {"message": "Hello, World!"}


# +++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
@app.get("/download/{filename}")
async def download_file(filename: str):
    """Download files"""

    loguru.logger.info(f"Received: GET /download/{filename}")
    try:
        # Decode the filename to handle spaces and special characters properly
        decoded_filename = urllib.parse.unquote(filename)

        debug_modee = 0
        if debug_modee == 1:
            loguru.logger.debug(f"Decoded filename: '{decoded_filename}'")

        # Path to look for files to be downloaded
        BASE_DIR2 = os.getcwd() + '/output/'

        debug_modee = 0
        if debug_modee == 1:
            loguru.logger.debug(f"Base directory: '{BASE_DIR2}'")

        # Path in filesystem for requested file
        file_path = os.path.join(BASE_DIR2, decoded_filename)  # Cambié de '/' a 'os.path.join'

        debug_modee = 0
        if debug_modee == 1:
            loguru.logger.debug(f"Full file path: '{file_path}'")

        # Check if file exists
        if os.path.exists(file_path):

            debug_modee = 0
            if debug_modee == 1:
                loguru.logger.debug(f"File found: {file_path}")

            return FileResponse(file_path, filename=decoded_filename)
        else:

            debug_modee = 0
            if debug_modee == 1:
                loguru.logger.debug(f"File not found: {file_path}")

            return JSONResponse(content={"error": f"File not found: {decoded_filename}"}, status_code=404)
    
    except Exception as e:
        logger.error(f"Error: {str(e)}")  # Usamos logger para los errores
        return JSONResponse(content={"error": str(e)}, status_code=500)
        

        
# +++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
@app.get("/export_excel_and_download", response_class=HTMLResponse)
async def get_export_excel_and_download():
    """Download files"""

    loguru.logger.info(f"Received: GET /export_excel_and_download")

    debug_modee = 0
    if debug_modee == 1:
        loguru.logger.debug("Exporting to Excel and Download...")

    # Execute for each view
    try:
        list_of_files = []
        for view in views:

            debug_modee = 0
            if debug_modee == 1:
                loguru.logger.debug(f"Generating Excel for Database view: {view}")
            # export_view_to_excel(view)
            generated_file = export_view_to_excel(view)

            debug_modee = 0
            if debug_modee == 1:
                loguru.logger.debug(f"generated_file: {generated_file}")

            list_of_files.append(generated_file)

            current_path = os.getcwd()
    
            debug_modee = 0
            if debug_modee == 1:
                loguru.logger.debug(f"Path: {current_path} ")
                loguru.logger.debug(f"List of files: {list_of_files} ")            
            
            if views == 'Table_VI_UOP_VIEW_4A_WITH_FE_EDCs_FOR_A_SHARED_UPF':
                generated_file_highlighted = generated_file.replace(".xlsx", "_highlighted.xlsx")
                list_of_files.append(generated_file_highlighted)

                debug_modee = 0
                if debug_modee == 1:
                    loguru.logger.debug(f"generated_file_highlighted: {generated_file_highlighted}")


        # Convert file names into downloadable URLs
        file_urls = [f"/download/{urllib.parse.quote(file)}" for file in list_of_files]

        debug_modee = 0
        if debug_modee == 1:
            loguru.logger.debug(f"file_urls: {file_urls}")
            
        # Replace /download/output/ by /download/
        file_urls = [url.replace('/download/output/', '/download/') for url in file_urls]

        debug_modee = 0
        if debug_modee == 1:
            loguru.logger.debug(f"file_urls: {file_urls}")
            loguru.logger.debug(f"BASE_DIR: '{BASE_DIR}'")

            for file_path in file_urls:

                if os.path.exists(full_file_path):
                    loguru.logger.debug(f"File existss: {full_file_path}")
                    
                else:
                    loguru.logger.debug(f"File does not existt: {full_file_path}")
                    

        return JSONResponse(content={"files": file_urls})

    except Exception as e:
        return JSONResponse(content={"error": str(e)}, status_code=500)
        

        
# +++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
@app.get("/export_excel", response_class=HTMLResponse)
async def get_export_excel():

    loguru.logger.info(f"Received: GET /export_excel")

    debug_modee = 1
    if debug_modee == 1:
        loguru.logger.debug(f"Exporting to Excel...")
        
    # Execute for each view
    try:
        list_of_files = []
        for view in views:

            debug_modee = 0
            if debug_modee == 1:
                loguru.logger.debug(f"Generating Excel for Database view: {view}")
            
            generated_file = export_view_to_excel(view)

        return JSONResponse(content={
            "message": "File generation OK",
            "path": BASE_DIR
        }, status_code=200)
            
    except Exception as e:
        return JSONResponse(content={"error": str(e)}, status_code=500)



# +++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
@app.get("/", response_class=HTMLResponse)
async def get_closest_device():

    loguru.logger.info(f"Received: GET /")

    loguru.logger.add(LOG_FILE, format="{time:YYYY-MM-DD HH:mm:ss.SSS} | {level} | {message}", rotation="1 day")

    debug_modee = 0
    if debug_modee == 1:
        loguru.logger.debug(f"random_value: {use_random_values}")
        
    if use_random_values == 0: ### We use original values

        if debug_modee == 1:
            loguru.logger.debug(f"Using default values from table Provisioned_devices")
            loguru.logger.debug(f"Updating table Provisioned_devices from Provisioned_devices_Original_values")
            
    else:

        if debug_modee == 1:
            loguru.logger.debug("Radnomize.......")
            
        all_data_html_1, all_data_html_2, all_data_html_3, all_data_html_4, my_summary_best_servers, my_fe_server_data, my_edge_server_data = get_random_value_in_db()
        
    debug_modee = 0


    # Fetch data from UOP_1_Shared_UPF_Edge_Far_Edge_DCs_location_info_v3
    data = execute_query_view(view_name)

    if not data:
        raise HTTPException(status_code=404, detail="No data found")    

    
    # Initialize variables for closest devices
    closest_devices = {
        "EdgeServer": None,
        "FEServer": None
    }

    max_score = {
        "EdgeServer": 0.0,
        "FEServer": 0.0
    }
    all_data_with_distances_all = []
    all_data_with_distances = []
    all_data_with_distances_edge = []




    ######################################################################################################
    ######################################################################################################
    # Process data to find the closest device 
    ######################################################################################################
    ######################################################################################################

    updated_data_ok = data
        
    new_data2 = []
    for row in updated_data_ok:
        try:
            # In table 'All Data from 'UOP_1_Shared_UPF_Edge_Far_Edge_DCs_location_info_v3' ( cells start with position 0 )
            value_type = row[4] # Role 
            value_latitude = float(row[17]) # Latitude
            value_longitude = float(row[18]) # Longitud
            latitude = value_latitude
            longitude = value_longitude
        except (ValueError, IndexError):
            continue  # Skip rows with invalid lat/long data

        # Calcuate distances (and latency) from SMU to each site ( FarEdge and Edge )
        distance = 0
        all_data_with_distances_all.append(row + (distance,))
        if value_type == 'FEServer':

            device_location = (latitude, longitude)
            
            # Calculate distance ------------------------------------------------------------------
            distance = round(haversine(my_location, device_location, unit=Unit.MILES), 2)
            
            ### Latency ---------------------------------------------------------------------------
            ### 200,000 km/s (124,000 miles per second)
            ### latency (ms) = [ distance in miles ] / [ 124 ]
#            latency =  round((distance / 124), 4)
#            latency =  round((distance / 200), 4)
            #latency =  round((distance * 3.4), 2) # theoretical carrier latency per kilometer is about 3.4μs for radio and 5μs for fiber

            # f(lt/μs)=distance (mi)*7.9   -- fiber
            # f(lt/μs)=distance (mi)*5.37  -- microwave
            latency =  round((distance * 5.37), 2) 


            ### Latency ---------------------------------------------------------------------------

            ### Score +++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++

            cpu_usage_value = row[6] # Column with CPU Usage
            ram_free_value = row[7] # Column with RAM free
            disk_free_value = row[8] # Column with Disk free 
            iops_value = row[9] # Column with IOPS 
            bandwith_value = row[10] # Column with Bandwidth 

            debug_modee = 0
            if debug_modee == 1:
                loguru.logger.debug("")
                loguru.logger.debug(f"Site '{row[1]}':")
     
            debug_modee = 0
            if debug_modee == 1:
                loguru.logger.debug("")
                loguru.logger.debug(f"- CPU score:")
                loguru.logger.debug(f"  cpu_weight     : {cpu_weight}")
                loguru.logger.debug(f"  cpu_usage_value: {cpu_usage_value}")
                loguru.logger.debug(f"  cpu_div        : {cpu_div}")
                loguru.logger.debug(f"  Formula        : cpu_weight * ( 1 - ( cpu_usage_value / cpu_div ) )")

            cpu_score = cpu_weight * ( 1 - ( cpu_usage_value / cpu_div ) )
            cpu_score_scaled = round(1 + 9 * cpu_score,2)
            

            debug_modee = 0
            if debug_modee == 1:
                loguru.logger.debug(f"  cpu_score        : {cpu_score}")
                loguru.logger.debug(f"  cpu_score_scaled : {cpu_score_scaled}")
                loguru.logger.debug("")


            debug_modee = 0
            if debug_modee == 1:
                loguru.logger.debug("")
                loguru.logger.debug(f"- RAM score:")
                loguru.logger.debug(f"  ram_weight     : {ram_weight}")
                loguru.logger.debug(f"  ram_free_value : {ram_free_value}")
                loguru.logger.debug(f"  ram_div        : {ram_div}")
                loguru.logger.debug(f"  Formula        : ram_weight * ( ram_free_value / ram_div ) ")

            ram_score = ram_weight * ( ram_free_value / ram_div )
            ram_score_scaled = round(1 + 9 * ram_score,2)

            debug_modee = 0
            if debug_modee == 1:
                loguru.logger.debug(f"  ram_score      : {ram_score}")
                loguru.logger.debug(f"  ram_score_scaled : {ram_score_scaled}")
                loguru.logger.debug("")

            debug_modee = 0
            if debug_modee == 1:
                loguru.logger.debug("")
                loguru.logger.debug(f"- Disk score:")
                loguru.logger.debug(f"  disk_weight     : {disk_weight}")
                loguru.logger.debug(f"  disk_free_value : {disk_free_value}")
                loguru.logger.debug(f"  disk_div        : {disk_div}")
                loguru.logger.debug(f"  Formula         : disk_weight * ( disk_free_value / disk_div ) ")

            disk_score = disk_weight * ( disk_free_value / disk_div )
            disk_score_scaled = round(1 + 9 * disk_score,2)

            debug_modee = 0
            if debug_modee == 1:
                loguru.logger.debug(f"  disk_score      : {disk_score}")
                loguru.logger.debug(f"  disk_score_scaled : {disk_score_scaled}")
                loguru.logger.debug("")
            
            debug_modee = 0
            if debug_modee == 1:
                loguru.logger.debug("")
                loguru.logger.debug(f"- IOPS score:")
                loguru.logger.debug(f"  iops_weight     : {iops_weight}")
                loguru.logger.debug(f"  iops_value      : {iops_value}")
                loguru.logger.debug(f"  iops_div        : {iops_div}")
                loguru.logger.debug(f"  iops_min_value  : {iops_min_value}")
                loguru.logger.debug(f"  iops_max_value  : {iops_max_value }")
                loguru.logger.debug(f"  Formula         : iops_weight * ( iops_value / iops_div ) ")
                loguru.logger.debug(f"  Formula2        : (iops - iops_min_value) / (iops_max_value - iops_min_value) ")

            #iops_score = iops_weight * ( iops_value / iops_div )
            # More iops is better
            # range is between 100,000–1,500,000
            #norm_iops = (iops - 100000) / 1400000  # Más IOPS es mejor
            iops_score = (iops_value - iops_min_value) / (iops_max_value - iops_min_value)  # Más IOPS es mejor
            iops_score_scaled = round(1 + 9 * iops_score,2)

            debug_modee = 0
            if debug_modee == 1:
                loguru.logger.debug(f"  iops_score      : {iops_score}")
                loguru.logger.debug(f"  iops_score_scaled : {iops_score_scaled}")
                loguru.logger.debug("")

            debug_modee = 0
            if debug_modee == 1:
                loguru.logger.debug("")
                loguru.logger.debug(f"- Bandwith score:")
                loguru.logger.debug(f"  bw_weight       : {bw_weight}")
                loguru.logger.debug(f"  bandwith_value  : {bandwith_value}")
                loguru.logger.debug(f"  bw_div          : {bw_div}")
                loguru.logger.debug(f"  bw_min_value    : {bw_min_value}")
                loguru.logger.debug(f"  bw_max_value    : {bw_max_value }")
                loguru.logger.debug(f"  Formula         : bw_weight * ( bandwith_value / bw_div ) ")
                loguru.logger.debug(f"  Formula2        : (bandwith_value - bw_min_value) / (bw_max_value - bw_min_value ) ")

            #bw_score = bw_weight * ( bandwith_value / bw_div )
            # More Bandwidth is better
            # range is 1GB - 25GB
            # norm_bw = (bw - 1) / 24  # Más BW es mejor 
            #norm_bw = (bw - 1) / 24  # Más BW es mejor
            bw_score = (bandwith_value - bw_min_value) / (bw_max_value - bw_min_value )
            bw_score_scaled = round(1 + 9 * bw_score,2)
            
            debug_modee = 0
            if debug_modee == 1:
                loguru.logger.debug(f"  bw_score        : {bw_score}")
                loguru.logger.debug(f"  bw_score_scaled : {bw_score_scaled}")
                loguru.logger.debug("")

            debug_modee = 0
            if debug_modee == 1:
                loguru.logger.debug("")
                loguru.logger.debug(f"- Distance score:")
                loguru.logger.debug(f"  distance_weight  : {distance_weight}")
                loguru.logger.debug(f"  distance_value   : {distance}")
                loguru.logger.debug(f"  distance_div     : {distance_div}")
                loguru.logger.debug(f"  Formula          : distance_weight * ( 1 - ( distance / distance_div ) )")


            ## We don't want to include the distance in final_score yet ######################
            ##distance_score = distance_weight * ( 1 - ( distance / distance_div ) )
            distance_score = 0

            debug_modee = 0
            if debug_modee == 1:
                loguru.logger.debug(f"  distance_score   : {distance_score}")
                loguru.logger.debug("")
            
            debug_modee = 0
            if debug_modee == 1:
                loguru.logger.debug("")
                loguru.logger.debug(f"- Latency score:")
                loguru.logger.debug(f"  latency_weight   : {latency_weight}")
                loguru.logger.debug(f"  latency_value    : {latency}")
                loguru.logger.debug(f"  latency_div      : {latency_div}")
                loguru.logger.debug(f"  latency_min_value: {latency_min_value}")
                loguru.logger.debug(f"  latency_max_value: {latency_max_value }")
                loguru.logger.debug(f"  Formula          : latency_weight * ( 1 - ( latency / latency_div ) )")
                loguru.logger.debug(f"  Formula2         : latency_weight * ( 1 - (latency - latency_min_value) / (latency_max_value - latency_min_value) )")

            #latency_score = latency_weight * ( 1 - ( latency / latency_div ) )
            latency_score = latency_weight * ( 1 - (latency - latency_min_value) / (latency_max_value - latency_min_value) )  # A menor latency mejor
            latency_score_scaled = round(1 + 9 * latency_score,2)

            debug_modee = 0
            if debug_modee == 1:
                loguru.logger.debug(f"  latency_score        : {latency_score}")
                loguru.logger.debug(f"  latency_score_scaled : {latency_score_scaled}")
                loguru.logger.debug("")
            
            debug_modee = 0
            if debug_modee == 1:
                loguru.logger.debug("")
                loguru.logger.debug(f"final_score = cpu_score ({cpu_score}) + ram_score ({ram_score}) + disk_score ({disk_score}) + iops_score ({iops_score}) + bw_score ({bw_score}) + latency_score ({latency_score}) ")

#            final_score = cpu_score + ram_score + disk_score + iops_score + bw_score + distance_score + latency_score
            final_score = cpu_score + ram_score + disk_score + iops_score + bw_score + latency_score
            final_score_scaled = round( ( cpu_score_scaled + ram_score_scaled + disk_score_scaled + iops_score_scaled + bw_score_scaled + latency_score_scaled ) / 6, 2)
            final_score = round(final_score, 4)
            final_score2 = round(final_score, 0)

            new_row = tuple(list(row) + [cpu_score_scaled] + [ram_score_scaled] + [disk_score_scaled] + [iops_score_scaled] + [bw_score_scaled] + [latency_score_scaled] + [final_score_scaled])
            new_data2.append(new_row)

            debug_modee = 0
            if debug_modee == 1:
                loguru.logger.debug(f"final_score = {final_score }")
                loguru.logger.debug(f"final_score2 = {final_score2 }")
                loguru.logger.debug(f"final_score2_scaled = {final_score_scaled }")
                loguru.logger.debug("")

            ffinal_score = final_score_scaled
            
            debug_modee = 0
            if debug_modee == 1:
                loguru.logger.debug(f">>> Final score: {ffinal_score}")


            ### Score +++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++

            all_data_with_distances.append(
                (ffinal_score,) + row + (distance,) + (latency,) + 
                (cpu_score_scaled,) + (ram_score_scaled,) + (disk_score_scaled,) + (iops_score_scaled,) + 
                (bw_score_scaled,) + (latency_score_scaled,) + 
                (final_score_scaled,)
                )

            
            # Determine type and update closest device if necessary
            value_device_type = row[4] ### Role
            device_type = value_device_type  
            
            debug_modee = 0
            if debug_modee == 1:
                loguru.logger.debug(f"value_device_type:{value_device_type}")
                loguru.logger.debug(f"device_type:{device_type}")
                loguru.logger.debug(closest_devices)

            if device_type in closest_devices:
                if ffinal_score > max_score[device_type]:
                    max_score[device_type] = ffinal_score
                    closest_devices[device_type] = (ffinal_score,) + row + (distance,) + (latency,) + (cpu_score_scaled,) + (ram_score_scaled,) + (disk_score_scaled,) + (iops_score_scaled,) + (bw_score_scaled,) + (latency_score_scaled,) + (final_score_scaled,)

    debug_modee = 0
    if debug_modee == 1:
        for item in new_data2:
            loguru.logger.debug(item)
        loguru.logger.debug("#####################################################")

    
###########################################################################
##
## At this point, we've choosen the FA server (cell site)
##
#############################################################################

    # Take coordinates of closest FarEdge server
    latitude_fe = closest_devices['FEServer'][18]
    longitude_fe = closest_devices['FEServer'][19]

    my_closest_fard_edge_is = (latitude_fe,longitude_fe)
    my_closest_fard_edge_name_is = closest_devices['FEServer'][14]
    
    # Process data to find the closest device EdgeSite and calculate distances
    new_data = []
    for row in updated_data_ok:
        try:
            value_type = row[4] # Role 
            value_latitude = float(row[17]) # Latitude 
            value_longitude = float(row[18]) # Longitude
            latitude = value_latitude
            longitude = value_longitude
        except (ValueError, IndexError):
            continue  # Skip rows with invalid lat/long data
        
        if value_type == 'EdgeServer':

            device_location = (latitude, longitude)

            # Calculate distance from my current EdgeSite to UTSW Medical Center
            distance_current_edge_site_to_utsw = round(haversine(device_location, utsw_coords, unit=Unit.MILES), 2)
            
            # Calculate distance from closes FarEdge site to my current EdgeSite
            #distance2_tmp = round(haversine(my_closest_fard_edge_is2, device_location2, unit=Unit.MILES), 2)
            distance_tmp  = round(haversine(my_closest_fard_edge_is, device_location, unit=Unit.MILES), 2)

            distance = round(distance_current_edge_site_to_utsw + distance_tmp,2)
            debug_modee = 0
            if debug_modee == 1:
                loguru.logger.debug(f"(A) Distance from selected FarEdge to this EdgeSite: {distance_tmp}miles")
                loguru.logger.debug(f"(B) Distance this EdgeSite to UTSW Medical Center  : {distance_current_edge_site_to_utsw}miles")
                loguru.logger.debug(f"Total distnace ( A + B )  : {distance}miles")
            
            # Calculate distance
#            distance = round(haversine(my_closest_fard_edge_is, device_location, unit=Unit.MILES), 2)

            ### Latency ---------------------------------------------------------------------------
            ### 200,000 km/s (124,000 miles per second)
            ### latency (ms) = [ distance in miles ] / [ 124 ]
#            latency =  round((distance / 124), 4)
#            latency =  round((distance / 200), 4)
            #latency =  round((distance * 5), 2) # theoretical carrier latency per kilometer is about 3.4μs for radio and 5μs for fiber

            # f(lt/μs)=distance (mi)*7.9   -- fiber
            # f(lt/μs)=distance (mi)*5.37  -- microwave
            latency =  round((distance * 7.9), 2) 
            
            ### Latency ---------------------------------------------------------------------------

            ### Score +++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++

                
            debug_modee = 0
            if debug_modee == 1:
                loguru.logger.debug("")
                loguru.logger.debug(f"- CPU score:")
                loguru.logger.debug(f"  cpu_weight     : {cpu_weight}")
                loguru.logger.debug(f"  cpu_usage_value: {cpu_usage_value}")
                loguru.logger.debug(f"  cpu_div        : {cpu_div}")
                loguru.logger.debug(f"  Formula        : cpu_weight * ( 1 - ( cpu_usage_value / cpu_div ) )")

            cpu_score = cpu_weight * ( 1 - ( cpu_usage_value / cpu_div ) )
            cpu_score_scaled = round(1 + 9 * cpu_score,2)

            debug_modee = 0
            if debug_modee == 1:
                loguru.logger.debug(f"  cpu_score        : {cpu_score}")
                loguru.logger.debug(f"  cpu_score_scaled : {cpu_score_scaled}")
                loguru.logger.debug("")

            debug_modee = 0
            if debug_modee == 1:
                loguru.logger.debug("")
                loguru.logger.debug(f"- RAM score:")
                loguru.logger.debug(f"  ram_weight     : {ram_weight}")
                loguru.logger.debug(f"  ram_free_value : {ram_free_value}")
                loguru.logger.debug(f"  ram_div        : {ram_div}")
                loguru.logger.debug(f"  Formula        : ram_weight * ( ram_free_value / ram_div ) ")

            ram_score = ram_weight * ( ram_free_value / ram_div )
            ram_score_scaled = round(1 + 9 * ram_score,2)

            debug_modee = 0
            if debug_modee == 1:
                loguru.logger.debug(f"  ram_score        : {ram_score}")
                loguru.logger.debug(f"  ram_score_scaled : {ram_score_scaled}")
                loguru.logger.debug("")

            debug_modee = 0
            if debug_modee == 1:
                loguru.logger.debug(f"- Disk score:")
                loguru.logger.debug(f"  disk_weight     : {disk_weight}")
                loguru.logger.debug(f"  disk_free_value : {disk_free_value}")
                loguru.logger.debug(f"  disk_div        : {disk_div}")
                loguru.logger.debug(f"  Formula         : disk_weight * ( disk_free_value / disk_div ) ")

            disk_score = disk_weight * ( disk_free_value / disk_div )
            disk_score_scaled = round(1 + 9 * disk_score,2)

            debug_modee = 0
            if debug_modee == 1:
                loguru.logger.debug(f"  disk_score        : {disk_score}")
                loguru.logger.debug(f"  disk_score_scaled : {disk_score_scaled}")
                loguru.logger.debug("")
            
            debug_modee = 0
            if debug_modee == 1:
                loguru.logger.debug("")
                loguru.logger.debug(f"- IOPS score:")
                loguru.logger.debug(f"  iops_weight     : {iops_weight}")
                loguru.logger.debug(f"  iops_value      : {iops_value}")
                loguru.logger.debug(f"  iops_div        : {iops_div}")
                loguru.logger.debug(f"  iops_min_value  : {iops_min_value}")
                loguru.logger.debug(f"  iops_max_value  : {iops_max_value }")
                loguru.logger.debug(f"  Formula         : iops_weight * ( iops_value / iops_div ) ")
                loguru.logger.debug(f"  Formula2        : (iops - iops_min_value) / (iops_max_value - iops_min_value) ")

            #iops_score = iops_weight * ( iops_value / iops_div )
            # More iops is better
            # range is between 100,000–1,500,000
            #norm_iops = (iops - 100000) / 1400000  # Más IOPS es mejor
            iops_score = (iops_value - iops_min_value) / (iops_max_value - iops_min_value)  # Más IOPS es mejor
            iops_score_scaled = round(1 + 9 * iops_score,2)
            
            debug_modee = 0
            if debug_modee == 1:
                loguru.logger.debug(f"  iops_score        : {iops_score}")
                loguru.logger.debug(f"  iops_score_scaled : {iops_score_scaled}")
                loguru.logger.debug("")

            debug_modee = 0
            if debug_modee == 1:
                loguru.logger.debug(f"- Bandwith score:")
                loguru.logger.debug(f"  bw_weight       : {bw_weight}")
                loguru.logger.debug(f"  bandwith_value  : {bandwith_value}")
                loguru.logger.debug(f"  bw_div          : {bw_div}")
                loguru.logger.debug(f"  bw_min_value    : {bw_min_value}")
                loguru.logger.debug(f"  bw_max_value    : {bw_max_value }")
                loguru.logger.debug(f"  Formula         : bw_weight * ( bandwith_value / bw_div ) ")
                loguru.logger.debug(f"  Formula2        : (bandwith_value - bw_min_value) / (bw_max_value - bw_min_value ) ")

            #bw_score = bw_weight * ( bandwith_value / bw_div )
            # More Bandwidth is better
            # range is 1GB - 25GB
            # norm_bw = (bw - 1) / 24  # Más BW es mejor 
            #norm_bw = (bw - 1) / 24  # Más BW es mejor
            bw_score = (bandwith_value - bw_min_value) / (bw_max_value - bw_min_value )
            bw_score_scaled = round(1 + 9 * bw_score,2)

            debug_modee = 0
            if debug_modee == 1:
                loguru.logger.debug(f"  bw_score        : {bw_score}")
                loguru.logger.debug(f"  bw_score_scaled : {bw_score_scaled}")
                loguru.logger.debug("")

            debug_modee = 0
            if debug_modee == 1:
                loguru.logger.debug("")
                loguru.logger.debug(f"- Distance score:")
                loguru.logger.debug(f"  distance_weight  : {distance_weight}")
                loguru.logger.debug(f"  distance_value   : {distance}")
                loguru.logger.debug(f"  distance_div     : {distance_div}")
                loguru.logger.debug(f"  Formula          : distance_weight * ( 1 - ( distance / distance_div ) )")
                loguru.logger.debug("")

# We don't want to include distance in score
#            distance_score = distance_weight * ( 1 - ( distance / distance_div ) )
            distance_score = 0

            debug_modee = 0
            if debug_modee == 1:
                loguru.logger.debug(f"  distance_score   : {distance_score}")
                loguru.logger.debug("")
            
            debug_modee = 0
            if debug_modee == 1:
                loguru.logger.debug("")
                loguru.logger.debug(f"- Latency score:")
                loguru.logger.debug(f"  latency_weight   : {latency_weight}")
                loguru.logger.debug(f"  latency_value    : {latency}")
                loguru.logger.debug(f"  latency_div      : {latency_div}")
                loguru.logger.debug(f"  latency_min_value: {latency_min_value}")
                loguru.logger.debug(f"  latency_max_value: {latency_max_value }")
                loguru.logger.debug(f"  latency_div      : {latency_div}")
                loguru.logger.debug(f"  Formula          : latency_weight * ( 1 - ( latency / latency_div ) )")
                loguru.logger.debug(f"  Formula2         : latency_weight * ( 1 - (latency - latency_min_value) / (latency_max_value - latency_min_value) )")

            #latency_score = 1 - (latency - latency_min_value) / (latency_max_value - latency_min_value)  # A menor latency mejor
            latency_score = latency_weight * ( 1 - (latency - latency_min_value) / (latency_max_value - latency_min_value) )  # A menor latency mejor
            latency_score_scaled = round(1 + 9 * latency_score,2)

            new_row = tuple(list(row) + [cpu_score_scaled] + [ram_score_scaled] + [disk_score_scaled] + [iops_score_scaled] + [bw_score_scaled] + [latency_score_scaled] + [final_score_scaled])
            new_data.append(new_row)
            
            debug_modee = 0
            if debug_modee == 1:
                loguru.logger.debug(f"  latency_score        : {latency_score}")
                loguru.logger.debug(f"  latency_score_scaled : {latency_score_scaled}")
                loguru.logger.debug("")
            
            debug_modee = 0
            if debug_modee == 1:
                loguru.logger.debug("")
                loguru.logger.debug(f"final_score = cpu_score ({cpu_score}) + ram_score ({ram_score}) + disk_score ({disk_score}) + iops_score ({iops_score}) + bw_score ({bw_score}) + distance_score ({distance_score}) + latency_score ({latency_score}) ")

            final_score = cpu_score + ram_score + disk_score + iops_score + bw_score + distance_score + latency_score
            final_score_scaled = round( ( cpu_score_scaled + ram_score_scaled + disk_score_scaled + iops_score_scaled + bw_score_scaled + latency_score_scaled ) / 6, 2)
            final_score = round(final_score, 4)
            final_score2 = round(final_score, 0)

            debug_modee = 0
            if debug_modee == 1:
                loguru.logger.debug(f"final_score = {final_score}")
                loguru.logger.debug(f"final_score2 = {final_score2}")
                loguru.logger.debug(f"final_score2_scaled = {final_score_scaled }")
                loguru.logger.debug("")
                loguru.logger.debug("")


            ffinal_score = final_score_scaled
            debug_modee = 0
            if debug_modee == 1:
                loguru.logger.debug(f">>> Final score: {ffinal_score}")

            ### Score +++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++

            all_data_with_distances_edge.append(
                (ffinal_score,) + row + (distance,) + (latency,) + 
                (cpu_score_scaled,) + (ram_score_scaled,) + (disk_score_scaled,) + (iops_score_scaled,) + 
                (bw_score_scaled,) + (latency_score_scaled,) + 
                (final_score_scaled,)
                )
            
            # Determine type and update closest device if necessary
            value_device_type = row[4] # Role
            device_type = value_device_type  


            debug_modee = 0
            if debug_modee == 1:
                loguru.logger.debug(f"value_device_type:{value_device_type}")
                loguru.logger.debug(f"device_type:{device_type}")
                loguru.logger.debug(closest_devices)

            if device_type in closest_devices:
                if ffinal_score > max_score[device_type]:
                    max_score[device_type] = ffinal_score
                    closest_devices[device_type] = (ffinal_score,) + row + (distance,) + (latency,) + (cpu_score_scaled,) + (ram_score_scaled,) + (disk_score_scaled,) + (iops_score_scaled,) + (bw_score_scaled,) + (latency_score_scaled,) + (final_score_scaled,)

#    print("--------------------~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")

    latitude_e = closest_devices['EdgeServer'][18]
    longitude_e = closest_devices['EdgeServer'][19]
    my_closest_edge_is = (latitude_e,longitude_e)
    my_closest_edge_name_is = closest_devices['EdgeServer'][14]
    
    # Render the HTML with tables and the map
    return render_gui(
        all_data_with_distances,
        closest_devices,
        my_closest_fard_edge_name_is,
        my_closest_edge_name_is,
        all_data_with_distances_edge,
        all_data_with_distances_all,
        all_data_html_1,
        all_data_html_2,
        all_data_html_3,
        all_data_html_4,
        my_summary_best_servers,
        my_fe_server_data,
        my_edge_server_data
    )





####################################################################
## 
## M A I N 
## 
####################################################################

# Dynamically get the filename of the current script (without extension)
current_script = Path(__file__).stem  # This will give you filename 

# Run Uvicorn with the provided or default parameters
if __name__ == "__main__":

    current_version = '1.0'

    loguru.logger.add(LOG_FILE, format="{time:YYYY-MM-DD HH:mm:ss.SSS} | {level} | {message}", rotation="1 day",level=MY_LOG_LEVEL)
    loguru.logger.info("")

    loguru.logger.info("+————————————————————————————————————————————+")
    loguru.logger.info(f"| UOCv{current_version} running at: http://{args.host}:{args.port}/ |")
    loguru.logger.info("+————————————————————————————————————————————+")
    loguru.logger.info("")

    loguru.logger.debug("Command line args:")
    loguru.logger.debug(f" + Host      : {args.host}")
    loguru.logger.debug(f" + Port      : {args.port}")
    loguru.logger.debug(f" + Reload    : {args.reload}")
    loguru.logger.debug(f" + Log level : {args.log_level}")
    loguru.logger.debug(f" + Log file  : {LOG_FILE}")
    loguru.logger.debug("")

    if use_random_values == 1: ### We use random values
        logger.warning(f"!! Using random values !! ")
        logger.warning("")

    uvicorn.run(f"{current_script}:app", host=args.host, port=args.port, reload=args.reload, log_level=args.log_level)


